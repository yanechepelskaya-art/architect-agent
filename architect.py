import requests
import time
import shutil
import os
import json
import sqlite3
import threading
from datetime import datetime, timedelta, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import ccxt
import websocket

API_KEY = "4197cf46-f0cd-475b-b0dc-31e42763c77c"
SECRET_KEY = "A088B1B1307E659FC85A8FF2A55A5899"
PASSPHRASE = "Kukuruza2026)"

TOKEN = "8900618226:AAGPVlCFCNSMiDYrv3DkUbeWorQWrdYti0Q"
CHAT_ID = "870512243"
CHANNEL_ID = "-1003920623687"

ACTIVE_ACTIVE_PORTFOLIO = {}
PASSIVE_PORTFOLIO = {}
WATCHLIST = ["BTC", "ETH", "SOL"]

ACTIVE_PORTFOLIO = {}
PASSIVE_PORTFOLIO = {}
PORTFOLIO = {
    "BTC": {"entry": 61900, "stop": 59500, "target": 68000, "amount": 0.01},
    "ETH": {"entry": 1550, "stop": 1450, "target": 2000, "amount": 0.1},
    "SOL": {"entry": 55, "stop": 48, "target": 82, "amount": 1.0},
    "DOGE": {"entry": 0.069, "stop": 0.060, "target": 0.085, "amount": 1000},
    "XRP": {"entry": 1.05, "stop": 0.90, "target": 1.30, "amount": 100},
    "SUI": {"entry": 0.68, "stop": 0.55, "target": 0.90, "amount": 50},
    "RENDER": {"entry": 1.42, "stop": 1.20, "target": 1.80, "amount": 50}
}

PASSIVE_ACTIVE_PORTFOLIO = {}
PASSIVE_PORTFOLIO = {}
WATCHLIST = ["BTC", "ETH", "SOL"]

ACTIVE_PORTFOLIO = {}
PASSIVE_PORTFOLIO = {}
PORTFOLIO = {"WLD": {"entry": 0.60, "stop": 0.50, "target": 0.70, "amount": 100}}
ACTIVE_PORTFOLIO = {}
PASSIVE_PORTFOLIO = {}
WATCHLIST = ["BTC", "ETH", "SOL"]

ACTIVE_PORTFOLIO = {}
PASSIVE_PORTFOLIO = {}
PORTFOLIO = {**ACTIVE_PORTFOLIO, **PASSIVE_PORTFOLIO}

BALANCE = 759.89
RISK_PERCENT = 2.0
TRAILING_PERCENT = 3.0
prev_price = None
prev_vol = None
last_phase = "Неизвестно"
last_update_id = 0
stop_warnings_sent = {}
target_warnings_sent = {}
PAPER_BALANCE = 10000.0
PAPER_POSITIONS = {}
ws_price = None
ws_vol = None
phase_start_time = datetime.now()
phase_history = []

def migrate_db():
    try:
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("PRAGMA table_info(trades)")
        cols = [col[1] for col in c.fetchall()]
        if "reason" not in cols:
            c.execute("ALTER TABLE trades ADD COLUMN reason TEXT")
            journal_event("migrate", "Added reason column to trades")
        if "stop_loss" not in cols:
            c.execute("ALTER TABLE trades ADD COLUMN stop_loss REAL")
        if "take_profit" not in cols:
            c.execute("ALTER TABLE trades ADD COLUMN take_profit REAL")
        conn.commit()
        conn.close()
    except:
        pass

def init_paper_db():
    migrate_db()
    conn = sqlite3.connect("paper_trades.db")
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS trades (id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT, coin TEXT, action TEXT, price REAL, amount REAL, pnl REAL, reason TEXT)")
    conn.commit(); conn.close()

def log_paper_trade(coin, action, price, amount, pnl=0, reason=""):
    conn = sqlite3.connect("paper_trades.db"); c = conn.cursor()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("INSERT INTO trades (timestamp, coin, action, price, amount, pnl, reason) VALUES (?,?,?,?,?,?,?)", (ts, coin, action, price, amount, pnl, reason))
    conn.commit(); conn.close()

def send_voice(text):
    try:
        url = f"https://api.telegram.org/bot{TOKEN}/sendVoice"
        import tempfile, os; from subprocess import run
        with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as f: f.write(text); f.flush(); tmp_path = f.name
        audio_path = tmp_path + ".m4a"
        run(["say", "-o", audio_path, "--data-format=alac", text], capture_output=True, timeout=10)
        with open(audio_path, 'rb') as af: requests.post(url, params={"chat_id": CHAT_ID}, files={"voice": af}, timeout=15)
        os.remove(tmp_path); os.remove(audio_path)
    except Exception as e:
        print(f"Update error: {e}")

def send_tg(text, reply_markup=None, to_channel=False):
    try:
        url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
        target = CHANNEL_ID if to_channel else CHAT_ID
        params = {"chat_id": target, "text": text, "parse_mode": "HTML"}
        if reply_markup and not to_channel: params["reply_markup"] = json.dumps(reply_markup)
        requests.get(url, params=params, timeout=10)
    except Exception as e:
        print(f"Update error: {e}")

def get_market_data_rest(symbol="BTC"):
    try:
        r = requests.get(f"https://www.okx.com/api/v5/market/ticker?instId={symbol}-USDT", timeout=5)
        d = r.json(); return float(d["data"][0]["last"]), float(d["data"][0]["vol24h"])
    except: return None, None

def get_price(symbol):
    if symbol == "BTC" and ws_price: return ws_price
    p, _ = get_market_data_rest(symbol); return p

def on_message(ws, message):
    global ws_price, ws_vol
    try:
        data = json.loads(message)
        if "data" in data and len(data["data"]) > 0: ws_price = float(data["data"][0]["last"]); ws_vol = float(data["data"][0]["vol24h"])
    except Exception as e:
        print(f"Update error: {e}")

def on_error(ws, error): pass
def on_close(ws, a, b): pass
def on_open(ws): ws.send(json.dumps({"op": "subscribe", "args": [{"channel": "tickers", "instId": "BTC-USDT"}]}))

def websocket_thread():
    while True:
        try:
            ws = websocket.WebSocketApp("wss://ws.okx.com:8443/ws/v5/public", on_open=on_open, on_message=on_message, on_error=on_error, on_close=on_close)
            ws.run_forever()
        except: time.sleep(5)

def export_to_excel():
    try:
        conn = sqlite3.connect("paper_trades.db"); c = conn.cursor()
        c.execute("SELECT * FROM trades ORDER BY id DESC"); rows = c.fetchall(); conn.close()
        wb = Workbook(); ws = wb.active; ws.title = "История сделок"
        for col, h in enumerate(["Дата", "Монета", "Действие", "Цена", "Объём", "PnL", "Причина"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for i, r in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=r[1]); ws.cell(row=i, column=2, value=r[2]); ws.cell(row=i, column=3, value=r[3])
            ws.cell(row=i, column=4, value=f"${r[4]:,.2f}"); ws.cell(row=i, column=5, value=f"{r[5]:.6f}")
            c = ws.cell(row=i, column=6, value=f"{r[6]:+,.2f}" if r[6] else "$0.00")
            c.font = Font(color="16A34A" if r[6] and r[6] >= 0 else "DC2626"); ws.cell(row=i, column=7, value=r[7])
        fn = f"trades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"; filepath = f"/Users/yananechepelskaya/Desktop/{fn}"
        wb.save(filepath)
        with open(filepath, 'rb') as f: requests.post(f"https://api.telegram.org/bot{TOKEN}/sendDocument", data={"chat_id": CHAT_ID}, files={"document": (fn, f)})
        import subprocess; subprocess.run(["open", filepath])
    except Exception as e: send_tg(f"Ошибка экспорта: {e}")

def predict():
    p, v = get_market_data_rest("BTC")
    if not p or not prev_price or not prev_vol:
        send_tg("⏳ Прогноз: жду данные")
        return
    ch = (p - prev_price) / prev_price * 100
    vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0

    # Определение сценария и фазы
    if ch > 0.5 and v > prev_vol:
        scenario, phase = "Бычий импульс", "Эманация"
        target, stop, conf, horizon = "$64,800–$65,500", "$63,800", "80%", "1–3 дня"
        action = "Держи лонги. Подтяни стопы."
        codex = "Кодекс: Можно наращивать позицию. Стоп двигай за ценой."
    elif ch > 0.5:
        scenario, phase = "Слабый рост", "Боковик"
        target, stop, conf, horizon = "$64,500–$64,800", "$63,500", "55%", "1–2 дня"
        action = "Не входи. Жди объёма."
        codex = "Кодекс: Без объёма нет импульса. Жди подтверждения."
    elif ch < -0.5 and v > prev_vol:
        scenario, phase = "Медвежий слив", "Сжатие"
        target, stop, conf, horizon = "$63,000–$62,700", "$64,200", "80%", "1–3 дня"
        action = "Проверь стопы. Не лови ножи."
        codex = "Кодекс: Защита депозита — приоритет. Фиксируй убытки рано."
    elif ch < -0.5:
        scenario, phase = "Слабое падение", "Боковик"
        target, stop, conf, horizon = "$63,300–$63,500", "$62,700", "55%", "1–2 дня"
        action = "Жди отскока."
        codex = "Кодекс: Не продавай на дне. Коррекция без объёма — шум."
    else:
        scenario, phase = "Боковик", "Боковик"
        target, stop, conf, horizon = "$63,500–$64,500", "$62,700", "40%", "—"
        action = "Не входи в середине."
        codex = "Кодекс: Терпение — позиция. Жди пробоя границ."

    # Сила сигнала
    strength = 5
    if abs(ch) > 2: strength += 2
    elif abs(ch) > 1: strength += 1
    if abs(vol_change) > 20: strength += 2
    elif abs(vol_change) > 10: strength += 1
    if ch > 0 and v > prev_vol: strength += 1
    elif ch < 0 and v > prev_vol: strength -= 1
    if strength > 10: strength = 10
    if strength < 1: strength = 1
    bar = "█" * strength + "░" * (10 - strength)

    reply = (
        f"🔮 <b>ПРОГНОЗ АРХИТЕКТОРА</b>\n"
        f"<code>══════════════════════</code>\n\n"
        f"₿ <b>BTC:</b> ${p:,.2f}\n"
        f"📊 <b>Изменение:</b> {ch:+.2f}%\n"
        f"📊 <b>Объём:</b> {vol_change:+.1f}%\n"
        f"⚡ <b>Сила сигнала:</b> {strength}/10 [{bar}]\n\n"
        f"📐 <b>Чертёж:</b> фаза — {phase}\n"
        f"🎯 <b>Сценарий:</b> {scenario}\n"
        f"🎯 <b>Цель:</b> {target}\n"
        f"🛑 <b>Стоп:</b> {stop}\n"
        f"📈 <b>Уверенность:</b> {conf}\n"
        f"⏳ <b>Горизонт:</b> {horizon}\n\n"
        f"💡 <b>Действие:</b> {action}\n"
        f"📋 {codex}"
    )
    send_tg(reply)

def backtest():
    send_tg("⏳ Backtest...")
    try:
        exchange = ccxt.okx(); since = int((datetime.now() - timedelta(days=30)).timestamp() * 1000)
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1d", since=since)
        balance = 10000.0; position = None; trades = 0; wins = 0
        for candle in ohlcv:
            close_price = candle[4]
            if position is None:
                if close_price < 64000: position = {"entry": close_price, "stop": close_price * 0.98, "target": close_price * 1.03}
            else:
                if close_price <= position["stop"]:
                    pnl = (close_price - position["entry"]) / position["entry"] * balance * 0.02; balance += pnl; trades += 1
                    if pnl > 0: wins += 1
                    position = None
                elif close_price >= position["target"]:
                    pnl = (close_price - position["entry"]) / position["entry"] * balance * 0.02; balance += pnl; trades += 1
                    if pnl > 0: wins += 1
                    position = None
        winrate = (wins / trades * 100) if trades > 0 else 0; profit = balance - 10000.0
        send_tg(f"📈 <b>BACKTEST</b>\n💰 Прибыль: <b>${profit:+,.2f}</b> ({profit/10000*100:+.2f}%)\n🎯 Винрейт: <b>{winrate:.1f}%</b>")
    except Exception as e:
        print(f"Update error: {e}")

def mystats():
    try:
        conn = sqlite3.connect("paper_trades.db"); c = conn.cursor()
        c.execute("SELECT pnl FROM trades WHERE pnl IS NOT NULL AND pnl != 0"); rows = c.fetchall(); conn.close()
        if not rows: send_tg("📊 Статистика пуста."); return
        pnls = [r[0] for r in rows]; total_pnl = sum(pnls); trades_count = len(pnls)
        wins = len([p for p in pnls if p > 0]); winrate = wins / trades_count * 100 if trades_count > 0 else 0
        best = max(pnls); worst = min(pnls)
        send_tg(f"📊 <b>СТАТИСТИКА</b>\n📋 Сделок: <b>{trades_count}</b>\n✅ Прибыльных: <b>{wins}</b>\n🎯 Винрейт: <b>{winrate:.1f}%</b>\n💰 PnL: <b>${total_pnl:+,.2f}</b>\n🟢 Лучшая: <b>${best:+,.2f}</b>\n🔴 Худшая: <b>${worst:+,.2f}</b>")
    except Exception as e:
        print(f"Update error: {e}")

def top_coins():
    try:
        conn = sqlite3.connect("paper_trades.db"); c = conn.cursor()
        c.execute("SELECT coin, SUM(pnl) FROM trades WHERE pnl IS NOT NULL GROUP BY coin ORDER BY SUM(pnl) DESC LIMIT 3"); rows = c.fetchall(); conn.close()
        if not rows: send_tg("🏆 Топ пуст."); return
        reply = "🏆 <b>ТОП МОНЕТ</b>\n"
        for i, r in enumerate(rows, 1): reply += f"{'🥇' if i==1 else ('🥈' if i==2 else '🥉')} <b>{r[0]}</b>: ${r[1]:+,.2f}\n"
        send_tg(reply)
    except Exception as e:
        print(f"Update error: {e}")

def check_stop_proximity():
    global stop_warnings_sent
    for coin, d in ACTIVE_PORTFOLIO.items():
        p = get_price(coin)
        if p:
            distance = (p - d["stop"]) / d["stop"] * 100
            if distance < 1.0 and (coin not in stop_warnings_sent or stop_warnings_sent[coin] < datetime.now() - timedelta(hours=1)):
                send_voice(f"Стоп-лосс близко! {coin}")
                send_tg(f"⚠️ <b>СТОП-ЛОСС БЛИЗКО!</b>\n{coin}: ${p:,.2f}, стоп ${d['stop']:,.2f}, запас {distance:.2f}%")
                stop_warnings_sent[coin] = datetime.now()

def check_target_proximity():
    global target_warnings_sent
    for coin, d in ACTIVE_PORTFOLIO.items():
        p = get_price(coin)
        if p:
            distance = (d["target"] - p) / p * 100
            if 0 < distance < 2.0 and (coin not in target_warnings_sent or target_warnings_sent[coin] < datetime.now() - timedelta(hours=1)):
                send_voice(f"Тейк-профит близко! {coin}")
                send_tg(f"🎯 <b>ТЕЙК-ПРОФИТ БЛИЗКО!</b>\n{coin}: ${p:,.2f}, цель ${d['target']:,.2f}, осталось {distance:.2f}%")
                target_warnings_sent[coin] = datetime.now()

def daily_channel_summary():
    p = get_price("BTC")
    # Статистика Paper Trading
    try:
        conn = sqlite3.connect("paper_trades.db"); c = conn.cursor()
        c.execute("SELECT pnl FROM trades WHERE pnl IS NOT NULL AND date(timestamp) = date('now')")
        today_rows = c.fetchall()
        c.execute("SELECT pnl FROM trades WHERE pnl IS NOT NULL")
        all_rows = c.fetchall(); conn.close()
        today_pnl = sum(r[0] for r in today_rows) if today_rows else 0
        today_trades = len(today_rows)
        total_pnl = sum(r[0] for r in all_rows) if all_rows else 0
        total_trades = len(all_rows)
        wins = len([r for r in all_rows if r[0] > 0])
        winrate = wins / total_trades * 100 if total_trades > 0 else 0
    except:
        today_pnl = today_trades = total_pnl = total_trades = winrate = 0
    # Фаза
    phase_icon = "📈" if "Эманация" in last_phase else ("📉" if "Сжатие" in last_phase else "📊")
    # Сентимент
    try:
        fg = requests.get("https://api.alternative.me/fng/", timeout=5).json()
        fg_value = fg["data"][0]["value"]
        fg_text = fg["data"][0]["value_classification"]
    except:
        fg_value, fg_text = "—", "нет данных"
    # Портфель
    total_portfolio_pnl = 0
    up_coins = down_coins = 0
    for coin, d in PORTFOLIO.items():
        cp = get_price(coin)
        if cp:
            pnl = (cp - d["entry"]) * d["amount"]
            total_portfolio_pnl += pnl
            if pnl >= 0: up_coins += 1
            else: down_coins += 1
    # Рекомендация на завтра
    if "Эманация" in last_phase:
        tomorrow = "Держи лонги. Подтягивай стопы. Можно наращивать позиции."
    elif "Сжатие" in last_phase:
        tomorrow = "Защити депозит. Проверь стоп-лоссы. Не лови ножи."
    else:
        tomorrow = "Жди пробоя. Не входи в середине диапазона."
    # Собираем отчёт
    summary = (
        f"🏰 <b>ДНЕВНОЙ ОТЧЁТ АРХИТЕКТОРА</b>\n"
        f"<code>══════════════════════</code>\n\n"
        f"📅 <b>{datetime.now().strftime('%d.%m.%Y')}</b>\n"
        f"{phase_icon} <b>Фаза:</b> {last_phase}\n"
        f"😱 <b>Страх/Жадность:</b> {fg_value}/100 ({fg_text})\n"
        f"₿ <b>BTC:</b> ${p:,.2f}" if p else "₿ BTC: нет данных" f"\n\n"
        f"<b>📊 СТАТИСТИКА ЗА ДЕНЬ</b>\n"
        f"Сделок сегодня: <b>{today_trades}</b>\n"
        f"PnL за день: <b>${today_pnl:+,.2f}</b>\n\n"
        f"<b>📊 ВСЕГО</b>\n"
        f"Сделок: <b>{total_trades}</b> | Винрейт: <b>{winrate:.0f}%</b>\n"
        f"Общий PnL: <b>${total_pnl:+,.2f}</b>\n\n"
        f"<b>💼 ПОРТФЕЛЬ</b>\n"
        f"💰 PnL: <b>${total_portfolio_pnl:+,.2f}</b>\n"
        f"🟢 В плюсе: {up_coins} | 🔴 В минусе: {down_coins}\n\n"
        f"<b>💡 НА ЗАВТРА</b>\n"
        f"{tomorrow}\n\n"
        f"<code>══════════════════════</code>\n"
        f"<i>Архитектор, Империя строится каждый день.</i> 🏰"
    )
    send_tg(summary, to_channel=True)

def process_updates():
    global last_update_id, last_phase
    try:
        if last_update_id == 0:
            # При первом запуске сбрасываем старые обновления
            r = requests.get(f"https://api.telegram.org/bot{TOKEN}/getUpdates", params={"offset": -1, "timeout": 1}, timeout=5).json()
            results = r.get("result", [])
            if results:
                last_update_id = results[-1]["update_id"]
        r = requests.get(f"https://api.telegram.org/bot{TOKEN}/getUpdates", params={"offset": last_update_id + 1, "timeout": 5}, timeout=10).json()
        for upd in r.get("result", []):
            last_update_id = upd["update_id"]
            msg = upd.get("message", {})
            t = msg.get("text", "")
            # Голосовое сообщение
            if "voice" in msg and not t:
                try:
                    file_id = msg["voice"]["file_id"]
                    file_info = requests.get(f"https://api.telegram.org/bot{TOKEN}/getFile?file_id={file_id}").json()
                    file_path = file_info["result"]["file_path"]
                    voice_url = f"https://api.telegram.org/file/bot{TOKEN}/{file_path}"
                    # Просто используем длительность как триггер (распознавание требует доп. API)
                    duration = msg["voice"]["duration"]
                    t = "/start"  # По умолчанию показываем Привет
                except:
                    t = ""
            # Голосовые команды (по длительности)
            if "voice" in upd.get("message", {}) and not upd["message"].get("text"):
                dur = upd["message"]["voice"].get("duration", 0)
                if dur <= 2:
                    t = "/sensor"
                elif dur <= 4:
                    t = "/predict"
                elif dur <= 6:
                    t = "/pulse"
                else:
                    t = "/strike"
            if t in ["/status", "📊 Статус"]:
                phase_icon = "📈" if "Эманация" in last_phase else ("📉" if "Сжатие" in last_phase else "📊")
                reply = f"🏰 <b>ПОРТФЕЛЬ</b>\n{phase_icon} {last_phase}\n<code>──</code>\n"
                for coin, d in PORTFOLIO.items():
                    p = get_price(coin)
                    if p:
                        pnl = (p - d["entry"]) / d["entry"] * 100
                        bar = "█" * min(int(abs(pnl)/10), 10) + "░" * max(10 - int(abs(pnl)/10), 0)
                        reply += f"{'🟢' if pnl>=0 else '🔴'} {coin}: ${p:,.2f} ({pnl:+.2f}%) [{bar}]\n"
                reply += f"💰 Баланс: ${BALANCE:,.2f}"
                send_tg(reply)
            elif t in ["/btc", "₿ BTC"]:
                p = get_price("BTC")
                if p: send_tg(f"₿ BTC: ${p:,.2f}")
            elif t in ["/risk", "⚠️ Риск"]:
                reply = "⚠️ <b>РИСК</b>\n"
                for coin, d in PORTFOLIO.items():
                    risk = d["amount"] * d["entry"] / BALANCE * 100
                    reply += f"{'🔴' if risk>RISK_PERCENT else '🟢'} {coin}: {risk:.1f}%\n"
                reply += f"\n📏 Лимит: {RISK_PERCENT}%"
                send_tg(reply)
            elif t in ["/sensor", "🧠 Сенсор"]:
                p, v = get_market_data_rest("BTC")
                if p and prev_price and prev_vol:
                    ch = (p - prev_price) / prev_price * 100
                    vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
                    reply = (
                        f"🧠 <b>СЕНСОР АРХИТЕКТОРА</b>\n"
                        f"<code>══════════════════════</code>\n\n"
                        f"₿ <b>BTC:</b> ${p:,.2f}\n"
                        f"📊 <b>Изменение:</b> {ch:+.2f}% за 5 мин\n"
                        f"📊 <b>Объём:</b> {vol_change:+.1f}% к прошлому замеру\n\n"
                    )
                    if ch > 0.5 and v > prev_vol:
                        reply += "🟢 <b>ЛЁГКОСТЬ</b>\nРынок растёт на объёме. Настоящий импульс.\n💡 <i>Кодекс: можно действовать. Стоп-лосс обязателен.</i>"
                    elif ch > 0.5:
                        reply += "🟡 <b>ТЯЖЕСТЬ (возможен ложный пробой)</b>\nЦена растёт, но объём падает.\n💡 <i>Кодекс: жди подтверждения. Не входи на эмоциях.</i>"
                    elif ch < -0.5 and v > prev_vol:
                        reply += "🔴 <b>ТЯЖЕСТЬ</b>\nРынок падает на объёме. Медведи активны.\n💡 <i>Кодекс: проверь стоп-лоссы. Будь готова.</i>"
                    elif ch < -0.5:
                        reply += "🟢 <b>ЛЁГКОСТЬ (возможна коррекция)</b>\nРынок падает без объёма. Продавцы слабы.\n💡 <i>Кодекс: наблюдай. Не продавай на дне.</i>"
                    else:
                        reply += "⚪ <b>НЕЙТРАЛЬНО</b>\nРынок в боковике. Нет явного сигнала.\n💡 <i>Кодекс: не входи в середине. Жди пробоя.</i>"
                    reply += f"\n\n📐 <i>Чертёж: текущая фаза — {last_phase}</i>"
                    send_tg(reply)
                else:
                    send_tg("⏳ <b>СЕНСОР АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\nСобираю данные... Нужно два замера.")
            elif t in ["/advice", "💡 Совет"]:
                p, v = get_market_data_rest("BTC")
                if not p or not prev_price or not prev_vol:
                    send_tg("💡 Совет: жду данные")
                    return
                ch = (p - prev_price) / prev_price * 100
                vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
                total_pnl = 0
                up_count = 0
                down_count = 0
                worst_coin = ("", 0)
                for coin, d in PORTFOLIO.items():
                    cp = get_price(coin)
                    if cp:
                        pnl = (cp - d["entry"]) * d["amount"]
                        total_pnl += pnl
                        pnl_pct = (cp - d["entry"]) / d["entry"] * 100
                        if pnl >= 0:
                            up_count += 1
                        else:
                            down_count += 1
                        if pnl_pct < worst_coin[1]:
                            worst_coin = (coin, pnl_pct)
                # Рыночная фаза
                if ch > 0.5 and v > prev_vol:
                    market_state = "📈 Эманация — рынок растёт на объёме"
                elif ch > 0.5:
                    market_state = "📈 Слабый рост — объём не подтверждает"
                elif ch < -0.5 and v > prev_vol:
                    market_state = "📉 Сжатие — рынок падает на объёме"
                elif ch < -0.5:
                    market_state = "📉 Слабое падение — объём низкий"
                else:
                    market_state = "📊 Боковик — консолидация"
                # Риски
                high_risk = [f"🔴 {coin}: {d['amount']*d['entry']/BALANCE*100:.1f}%" for coin, d in PORTFOLIO.items() if get_price(coin) and d['amount']*d['entry']/BALANCE*100 > RISK_PERCENT]
                # Рекомендация + Кодекс
                if ch > 0.5 and v > prev_vol and total_pnl > 10:
                    action = "✅ Держи позиции. Подтяни стоп-лоссы в плюс."
                    codex = "Кодекс: В Эманации можно наращивать. Но стоп двигай за ценой. Не жадничай."
                elif ch > 0.5 and total_pnl > 10:
                    action = "⚠️ Рост есть, но без объёма. Зафиксируй часть прибыли."
                    codex = "Кодекс: Слабый рост — ловушка. Лучше взять мало, чем потерять много."
                elif ch < -0.5 and v > prev_vol and total_pnl < -10:
                    action = "🔴 Рынок падает на объёме. Проверь стопы. Не усредняй."
                    codex = "Кодекс: В Сжатии защита депозита — приоритет. Фиксируй убытки рано."
                elif ch < -0.5 and total_pnl < -10:
                    action = "🟡 Падение без объёма. Возможен отскок. Не паникуй."
                    codex = "Кодекс: Не продавай на дне. Коррекция без объёма — шум. Жди."
                elif down_count > up_count:
                    action = "⚠️ Портфель в минусе. Проверь худшую позицию."
                    codex = f"Кодекс: Худшая монета — {worst_coin[0]} ({worst_coin[1]:+.1f}%). Режь убытки, не дай им расти."
                else:
                    action = "⚪ Жди. Рынок не даёт сильного сигнала."
                    codex = "Кодекс: Терпение — позиция. Не входи в середине диапазона."
                reply = (
                    f"💡 <b>СОВЕТ АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"{market_state}\n"
                    f"📊 Объём: {vol_change:+.1f}%\n\n"
                    f"💰 <b>Портфель:</b> ${total_pnl:+,.2f}\n"
                    f"🟢 В плюсе: {up_count} | 🔴 В минусе: {down_count}\n\n"
                    f"⚠️ <b>Риски:</b>\n"
                )
                if high_risk:
                    reply += "\n".join(high_risk) + "\n"
                else:
                    reply += "🟢 Всё в норме.\n"
                reply += (
                    f"\n💡 <b>Действие:</b> {action}\n"
                    f"📋 {codex}"
                )
                send_tg(reply)
            elif t in ["/blueprint", "📐 Чертёж"]:
                p = get_price("BTC")
                if not p: send_tg("📐 Чертёж: нет данных"); return
                duration = datetime.now() - phase_start_time
                hours = duration.seconds // 3600
                minutes = (duration.seconds % 3600) // 60
                days = duration.days
                if "Эманация" in last_phase:
                    stage, next_phase, risk, action, signal = "Рост", "Сжатие", "Средний", "Держи. Подтягивай стопы.", "🟢"
                    codex = "Кодекс: Можно наращивать позицию. Стоп двигай за ценой. Не жадничай."
                    levels = "Сопротивление: $65,000 / $65,500"
                elif "Сжатие" in last_phase:
                    stage, next_phase, risk, action, signal = "Спад", "Эманация", "Повышенный", "Проверь стопы. Ищи точку входа.", "🔴"
                    codex = "Кодекс: Защита депозита — приоритет. Фиксируй убытки рано. Не усредняй."
                    levels = "Поддержка: $62,700 / $62,000"
                else:
                    stage, next_phase, risk, action, signal = "Нейтрально", "Пробой", "Низкий", "Жди. Лимитные ордера.", "⚪"
                    codex = "Кодекс: Терпение — позиция. Не входи в середине диапазона. Жди пробоя."
                    levels = "Диапазон: $63,500–$64,500"
                dur_str = f"{days}д {hours}ч {minutes}мин" if days > 0 else f"{hours}ч {minutes}мин"
                dur_bar = "⏳" + "█" * min(int(hours / 2), 10) + "░" * max(10 - int(hours / 2), 0)
                reply = (
                    f"📐 <b>ЧЕРТЁЖ АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"{signal} <b>Фаза: {last_phase}</b>\n"
                    f"📊 Стадия: {stage}\n"
                    f"⏱ Длительность: {dur_str} {dur_bar}\n\n"
                    f"🔮 <b>Далее:</b> {next_phase}\n"
                    f"⚠️ <b>Риск:</b> {risk}\n"
                    f"📐 <b>Уровни:</b> {levels}\n\n"
                    f"💡 <b>Действие:</b> {action}\n"
                    f"📋 {codex}\n\n"
                    f"₿ BTC: <b>${p:,.2f}</b>"
                )
                send_tg(reply)
            elif t in ["/energy", "⚡ Энергия"]:
                p, v = get_market_data_rest("BTC")
                if not p or not prev_price or not prev_vol:
                    send_tg("⚡ <b>ЭНЕРГИЯ АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\n⏳ Собираю данные... Нужно два замера.")
                    return
                ch = (p - prev_price) / prev_price * 100
                vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
                strength = 5
                if abs(ch) > 2: strength += 2
                elif abs(ch) > 1: strength += 1
                if abs(vol_change) > 20: strength += 2
                elif abs(vol_change) > 10: strength += 1
                if ch > 0 and v > prev_vol: strength += 1
                elif ch < 0 and v > prev_vol: strength -= 1
                if strength > 10: strength = 10
                if strength < 1: strength = 1
                bar = "█" * strength + "░" * (10 - strength)
                if strength >= 7:
                    level = "🔥 ВЫСОКАЯ"
                    character = "Рынок активен. Крупный игрок в движении. Возможны сильные импульсы."
                    warning = "Будь готова к резким движениям. Стопы обязательны."
                    imperative = "🚫 ИМПЕРАТИВ: Можно входить. Размер позиции — 2% депозита. Стоп держи близко."
                elif strength >= 4:
                    level = "⚡ СРЕДНЯЯ"
                    character = "Рынок в рабочем режиме. Движения плавные, предсказуемые."
                    warning = "Следуй Кодексу. Не форсируй сделки без сигнала."
                    imperative = "⚠️ ИМПЕРАТИВ: Стандартный размер — 1% депозита. Жди подтверждения."
                else:
                    level = "💤 НИЗКАЯ"
                    character = "Рынок спит. Нет явного интереса. Объёмы низкие."
                    warning = "Не входи в рынок без крайней необходимости. Жди пробуждения."
                    imperative = "🚫 ИМПЕРАТИВ: Новые сделки запрещены. Энергии нет."
                reply = (
                    f"⚡ <b>ЭНЕРГИЯ АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n\n"
                    f"<b>📊 СИЛА СИГНАЛА</b>\n"
                    f"<b>{strength}/10</b> [{bar}]\n"
                    f"Уровень: {level}\n\n"
                    f"<b>🔍 ХАРАКТЕР РЫНКА</b>\n"
                    f"{character}\n\n"
                    f"<b>📊 ДАННЫЕ</b>\n"
                    f"₿ BTC: <b>${p:,.2f}</b>\n"
                    f"Изменение: {ch:+.2f}%\n"
                    f"Объём: {vol_change:+.1f}%\n\n"
                    f"<b>📋 КОДЕКС</b>\n"
                    f"{warning}\n\n"
                    f"{imperative}\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"<i>Архитектор, энергия — это топливо. Используй её с умом.</i>"
                )
                send_tg(reply)
            elif t in ["/shadow", "👁 Тень"]:
                p, v = get_market_data_rest("BTC")
                if not p or not prev_price or not prev_vol:
                    send_tg("👁 <b>ТЕНЬ АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\n⏳ Собираю данные... Нужно два замера.")
                    return
                ch = (p - prev_price) / prev_price * 100
                vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
                reply = (
                    f"👁 <b>ТЕНЬ АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n"
                    f"₿ <b>BTC:</b> ${p:,.2f}\n"
                    f"📊 <b>Изменение:</b> {ch:+.2f}%\n"
                    f"📊 <b>Объём:</b> {vol_change:+.1f}%\n\n"
                    f"<b>🔍 ДЕТЕКТОР МАНИПУЛЯЦИЙ</b>\n"
                )
                warnings = []
                if abs(ch) > 1 and abs(vol_change) < 10:
                    warnings.append(("⚠️ ЛОЖНЫЙ ПРОБОЙ (STOP HUNT)", "Цена движется сильно, но объём низкий. Маркет-мейкер выбивает стопы толпы.", "Не двигай стоп-лосс. Не входи на резком движении без объёма. Жди возврата цены."))
                if abs(ch) < 0.3 and abs(vol_change) > 20:
                    warnings.append(("🔄 НАКОПЛЕНИЕ", "Объём растёт, а цена стоит на месте. Крупный игрок набирает позицию.", "Не торопись. Возможен сильный импульс после накопления. Будь готова."))
                if ch < -1 and v > prev_vol * 1.5:
                    warnings.append(("🔴 ПАНИКА / КАПИТУЛЯЦИЯ", "Резкое падение на высоком объёме. Толпа сливает активы.", "Не поддавайся панике. Проверь стопы. Возможен отскок после капитуляции."))
                if ch > 1 and v > prev_vol * 1.5:
                    warnings.append(("🟢 FOMO / ЭЙФОРИЯ", "Резкий рост на высоком объёме. Толпа заходит на эмоциях.", "Не входи на пике. Жди коррекции. Не жадничай — фиксируй прибыль."))
                if warnings:
                    reply += f"⚠️ <b>ОБНАРУЖЕНО СИГНАЛОВ: {len(warnings)}</b>\n\n"
                    for i, (title, desc, action) in enumerate(warnings, 1):
                        reply += f"<b>{i}. {title}</b>\n"
                        reply += f"🔍 <b>Причина:</b> {desc}\n"
                        reply += f"📋 <b>Кодекс:</b> {action}\n\n"
                else:
                    reply += "✅ <b>Манипуляций не обнаружено.</b>\nРынок ведёт себя естественно.\n\n"
                # Императив
                if len(warnings) >= 2:
                    reply += "🚫 <b>ИМПЕРАТИВ:</b> Новые сделки запрещены. Рынок опасен."
                elif len(warnings) == 1:
                    reply += "⚠️ <b>ИМПЕРАТИВ:</b> Размер позиций — половина стандартного."
                else:
                    reply += "💚 <b>ИМПЕРАТИВ:</b> Можно действовать по сигналу Чертёжа."
                reply += f"\n\n<code>══════════════════════</code>\n<i>Архитектор, тень — это иллюзия. Ты видишь истину.</i>"
                send_tg(reply)
            elif t in ["/breath", "🌬 Дыхание"]:
                if not prev_price:
                    send_tg("🌬 <b>ДЫХАНИЕ</b>\nСобираю данные..."); return
                p = get_price("BTC")
                duration = datetime.now() - phase_start_time
                hours = duration.seconds // 3600
                minutes = (duration.seconds % 3600) // 60
                reply = f"🌬 <b>ДЫХАНИЕ РЫНКА: АНАЛИЗ ЦИКЛА</b>\n"
                reply += f"<code>══════════════════════════</code>\n\n"
                reply += f"<b>📈 ТЕКУЩАЯ ФАЗА</b>\n"
                reply += f"{last_phase}\n"
                reply += f"Длительность: <b>{hours} ч {minutes} мин</b>\n\n"
                if "Эманация" in last_phase:
                    reply += f"<b>🔍 ХАРАКТЕР ФАЗЫ</b>\n"
                    reply += f"Рынок в фазе роста. Энергия расширения.\n\n"
                    if hours >= 12:
                        reply += f"<b>⚠️ СТАДИЯ:</b> Зрелая.\n"
                        reply += f"Фаза Эманации длится более 12 часов. Вероятность перехода в Сжатие — <b>высокая</b>.\n\n"
                        reply += f"<b>💡 КОДЕКС:</b>\n"
                        reply += f"• Готовься фиксировать прибыль.\n"
                        reply += f"• Подтяни стопы в безубыток.\n"
                        reply += f"• Не открывай новые лонги на пике."
                    elif hours >= 6:
                        reply += f"<b>⚠️ СТАДИЯ:</b> Развитая.\n"
                        reply += f"Фаза Эманации в разгаре. Ещё есть потенциал для роста.\n\n"
                        reply += f"<b>💡 КОДЕКС:</b>\n"
                        reply += f"• Держи позиции.\n"
                        reply += f"• Можно частично фиксировать прибыль.\n"
                        reply += f"• Следи за объёмами — падение объёма = сигнал к выходу."
                    else:
                        reply += f"<b>🟢 СТАДИЯ:</b> Молодая.\n"
                        reply += f"Фаза только началась. Импульс свежий.\n\n"
                        reply += f"<b>💡 КОДЕКС:</b>\n"
                        reply += f"• Можно входить в лонг.\n"
                        reply += f"• Стоп под уровень последнего Сжатия."
                elif "Сжатие" in last_phase:
                    reply += f"<b>🔍 ХАРАКТЕР ФАЗЫ</b>\n"
                    reply += f"Рынок в фазе сжатия. Накопление энергии.\n\n"
                    if hours >= 12:
                        reply += f"<b>🟢 СТАДИЯ:</b> Зрелая.\n"
                        reply += f"Сжатие затянулось. Вероятность разворота в Эманацию — <b>высокая</b>.\n\n"
                        reply += f"<b>💡 КОДЕКС:</b>\n"
                        reply += f"• Ищи точку входа.\n"
                        reply += f"• Жди подтверждения — рост на объёме.\n"
                        reply += f"• Не входи раньше времени."
                    elif hours >= 6:
                        reply += f"<b>⚠️ СТАДИЯ:</b> Развитая.\n"
                        reply += f"Сжатие продолжается. Возможно ещё одно движение вниз.\n\n"
                        reply += f"<b>💡 КОДЕКС:</b>\n"
                        reply += f"• Не торопись с входом.\n"
                        reply += f"• Проверь стоп-лоссы.\n"
                        reply += f"• Следи за объёмами — рост объёма при падении = капитуляция."
                    else:
                        reply += f"<b>🔴 СТАДИЯ:</b> Молодая.\n"
                        reply += f"Сжатие только началось. Возможно углубление.\n\n"
                        reply += f"<b>💡 КОДЕКС:</b>\n"
                        reply += f"• Не лови падающие ножи.\n"
                        reply += f"• Дай рынку устаканиться."
                else:
                    reply += f"<b>🔍 ХАРАКТЕР ФАЗЫ</b>\n"
                    reply += f"Рынок в боковике. Неопределённость.\n\n"
                    reply += f"<b>💡 КОДЕКС:</b>\n"
                    reply += f"• Не входи в середине диапазона.\n"
                    reply += f"• Используй лимитные ордера у границ.\n"
                    reply += f"• Жди пробоя."
                if p:
                    reply += f"\n<b>₿ BTC:</b> ${p:,.2f}\n"
                reply += f"\n<code>══════════════════════════</code>\n"
                reply += f"<i>Архитектор, рынок дышит циклами. Вдох — Сжатие, выдох — Эманация.</i>"
                send_tg(reply)
            elif t in ["/pulse", "💓 Пульс"]:
                total_coins = up_coins = down_coins = 0; total_pnl = 0
                best_coin = ("", -999); worst_coin = ("", 999)
                for coin, d in PORTFOLIO.items():
                    p = get_price(coin)
                    if p:
                        total_coins += 1; pnl = (p - d["entry"]) / d["entry"] * 100; total_pnl += pnl
                        if pnl >= 0: up_coins += 1
                        else: down_coins += 1
                        if pnl > best_coin[1]: best_coin = (coin, pnl)
                        if pnl < worst_coin[1]: worst_coin = (coin, pnl)
                if total_coins == 0:
                    send_tg("💓 <b>ПУЛЬС АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\nНет данных.")
                    return
                health = up_coins / total_coins * 100; avg_pnl = total_pnl / total_coins
                if health >= 70:
                    status = "💚 Здоров"
                    advice = "Портфель в отличной форме. Можно масштабировать позиции."
                    imperative = "💚 ИМПЕРАТИВ: Атака разрешена. Размер позиций — 2% депозита."
                elif health >= 40:
                    status = "💛 Стабилен"
                    advice = "Портфель в норме. Следи за убыточными позициями."
                    imperative = "⚠️ ИМПЕРАТИВ: Защита. Размер позиций — 1% депозита."
                else:
                    status = "❤️ Требует внимания"
                    advice = "Слишком много убыточных позиций. Проверь стоп-лоссы."
                    imperative = "🚫 ИМПЕРАТИВ: Новые сделки запрещены. Зафиксируй убытки."
                reply = (
                    f"💓 <b>ПУЛЬС АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n\n"
                    f"<b>📊 ОБЩАЯ СТАТИСТИКА</b>\n"
                    f"Монет в портфеле: <b>{total_coins}</b>\n"
                    f"🟢 В плюсе: <b>{up_coins}</b>\n"
                    f"🔴 В минусе: <b>{down_coins}</b>\n"
                    f"Средний PnL: <b>{avg_pnl:+.2f}%</b>\n\n"
                    f"<b>💪 ЗДОРОВЬЕ ПОРТФЕЛЯ</b>\n"
                    f"<b>{health:.0f}%</b> — {status}\n"
                    f"[{'█' * int(health/10)}{'░' * (10 - int(health/10))}]\n\n"
                )
                if best_coin[0]:
                    reply += f"<b>🏆 ЛУЧШАЯ МОНЕТА</b>\n🟢 {best_coin[0]}: <b>{best_coin[1]:+.2f}%</b>\n\n"
                if worst_coin[0]:
                    reply += f"<b>⚠️ ХУДШАЯ МОНЕТА</b>\n🔴 {worst_coin[0]}: <b>{worst_coin[1]:+.2f}%</b>\n\n"
                reply += (
                    f"<b>📋 КОДЕКС</b>\n{advice}\n\n"
                    f"{imperative}\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"<i>Архитектор, пульс — это ритм твоей Империи. Держи его ровным.</i>"
                )
                send_tg(reply)
            elif t in ["/levels", "🗺 Уровни"]:
                p = get_price("BTC")
                if not p:
                    send_tg("🗺 <b>УРОВНИ АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\nНет данных.")
                    return
                ma20 = round(p * 0.98, 2)
                ma60 = round(p * 1.05, 2)
                ma120 = round(p * 1.08, 2)
                psych_up = (int(p/1000)+1)*1000
                psych_down = int(p/1000)*1000
                if p > ma60:
                    zone = "🟢 БЫЧЬЯ ЗОНА"
                    zone_desc = "Цена выше MA60. Рынок в восходящем тренде."
                    action = "Можно удерживать лонги. Стоп под MA60."
                    imperative = "💚 ИМПЕРАТИВ: Можно действовать. Размер позиций — 2% депозита."
                elif p > ma20:
                    zone = "⚪ НЕЙТРАЛЬНАЯ ЗОНА"
                    zone_desc = "Цена между MA20 и MA60. Неопределённость."
                    action = "Жди пробоя. Не входи в середине."
                    imperative = "⚠️ ИМПЕРАТИВ: Защита. Размер позиций — 1% депозита."
                else:
                    zone = "🔴 МЕДВЕЖЬЯ ЗОНА"
                    zone_desc = "Цена ниже MA20. Нисходящий тренд."
                    action = "Проверь стопы. Не лови падающие ножи."
                    imperative = "🚫 ИМПЕРАТИВ: Новые сделки запрещены. Защищай депозит."
                reply = (
                    f"🗺 <b>УРОВНИ АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n"
                    f"₿ <b>BTC:</b> ${p:,.2f}\n"
                    f"📍 <b>Зона:</b> {zone}\n"
                    f"{zone_desc}\n\n"
                    f"<b>🔴 СОПРОТИВЛЕНИЕ</b>\n"
                    f"• MA20: <b>${ma20:,.2f}</b> — ближайшее\n"
                    f"• MA60: <b>${ma60:,.2f}</b> — среднее\n"
                    f"• MA120: <b>${ma120:,.2f}</b> — сильное\n"
                    f"• Психологический: <b>${psych_up:,.2f}</b>\n\n"
                    f"<b>🟢 ПОДДЕРЖКА</b>\n"
                    f"• MA20: <b>${ma20:,.2f}</b> — ближайшая\n"
                    f"• Психологический: <b>${psych_down:,.2f}</b>\n"
                    f"• Стоп-лосс: <b>$62,700</b> — твоя броня\n\n"
                    f"<b>💡 ДЕЙСТВИЕ</b>\n{action}\n\n"
                    f"{imperative}\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"<i>Архитектор, уровни — это карта. Ты знаешь, где вход, а где выход.</i>"
                )
                send_tg(reply)
            elif t in ["/mirror", "🪞 Зеркало"]:
                try:
                    conn = sqlite3.connect("paper_trades.db"); c = conn.cursor()
                    c.execute("SELECT action, pnl, reason FROM trades WHERE pnl IS NOT NULL ORDER BY id DESC LIMIT 10")
                    rows = c.fetchall(); conn.close()
                    if not rows:
                        send_tg("🪞 <b>ЗЕРКАЛО АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\nНет данных. Агент набирает статистику.")
                        return
                    wins = len([r for r in rows if r[1] > 0])
                    losses = len([r for r in rows if r[1] < 0])
                    total_pnl = sum([r[1] for r in rows])
                    winrate = wins / len(rows) * 100 if len(rows) > 0 else 0
                    avg_win = sum([r[1] for r in rows if r[1] > 0]) / wins if wins > 0 else 0
                    avg_loss = sum([r[1] for r in rows if r[1] < 0]) / losses if losses > 0 else 0

                    reply = (
                        f"🪞 <b>ЗЕРКАЛО АРХИТЕКТОРА</b>\n"
                        f"<code>══════════════════════</code>\n\n"
                        f"📐 <b>Чертёж:</b> фаза — {last_phase}\n\n"
                        f"<b>📊 СТАТИСТИКА (последние 10 сделок)</b>\n"
                        f"Всего: <b>{len(rows)}</b>\n"
                        f"✅ Прибыльных: <b>{wins}</b>\n"
                        f"❌ Убыточных: <b>{losses}</b>\n"
                        f"🎯 Винрейт: <b>{winrate:.0f}%</b>\n"
                        f"💰 Общий PnL: <b>${total_pnl:+,.2f}</b>\n\n"
                    )

                    if wins > 0 and losses > 0:
                        reply += (
                            f"<b>📈 СРЕДНИЕ ЗНАЧЕНИЯ</b>\n"
                            f"Средняя прибыль: <b>${avg_win:+,.2f}</b>\n"
                            f"Средний убыток: <b>${avg_loss:+,.2f}</b>\n"
                        )
                        if avg_win > abs(avg_loss):
                            reply += f"✅ Прибыль больше убытка — <b>хороший риск-менеджмент</b>.\n\n"
                        else:
                            reply += f"⚠️ Убыток больше прибыли — <b>проверь соотношение риск/прибыль</b>.\n\n"

                    # Кодекс и Императив
                    if winrate >= 70:
                        reply += "🟢 <b>Отличный винрейт!</b> Твой Сенсор точен.\n"
                        codex = "Кодекс: Продолжай в том же духе. Масштабируй позиции."
                        imperative = "💚 ИМПЕРАТИВ: Атака разрешена. Размер позиций — 2% депозита."
                    elif winrate >= 50:
                        reply += "💛 <b>Стабильно.</b> Ты на правильном пути.\n"
                        codex = "Кодекс: Есть куда расти. Анализируй убыточные сделки."
                        imperative = "⚠️ ИМПЕРАТИВ: Защита. Размер позиций — 1% депозита."
                    else:
                        reply += "⚠️ <b>Винрейт ниже 50%.</b>\n• Не входишь ли на эмоциях?\n• Соблюдаешь ли стоп-лосс?\n• Не входишь ли в середине диапазона?\n"
                        codex = "Кодекс: Пересмотри стратегию. Тренируйся на Paper Trading."
                        imperative = "🚫 ИМПЕРАТИВ: Реальные сделки запрещены. Только симуляция."

                    if losses > 0:
                        stop_losses = len([r for r in rows if r[1] < 0 and "Стоп" in (r[2] or "")])
                        if stop_losses > 0:
                            reply += f"\n📋 Из {losses} убыточных — <b>{stop_losses}</b> по стоп-лоссу. Это плата за дисциплину.\n"

                    reply += (
                        f"\n<code>══════════════════════</code>\n"
                        f"📋 {codex}\n"
                        f"{imperative}\n\n"
                        f"<i>Архитектор, зеркало не судит. Оно показывает правду. Используй её.</i>"
                    )
                    send_tg(reply)
                except:
                    send_tg("🪞 <b>ЗЕРКАЛО АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\nОшибка доступа к данным.")
            elif t in ["/beacon", "🏮 Маяк"]:
                p = get_price("BTC")
                if not p:
                    send_tg("🏮 <b>МАЯК АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\nНет данных.")
                    return

                if p > 64000:
                    trend = "📈 Восходящий"
                    target = "$65,000 → $68,000"
                    stop = "$63,500"
                    action = "Держи лонги. Подтягивай стопы в безубыток. Частично фиксируй прибыль на уровнях."
                    risk = "Ниже среднего"
                    imperative = "💚 ИМПЕРАТИВ: Можно наращивать позиции. Размер — 2% депозита."
                elif p > 62000:
                    trend = "📊 Боковик с восходящим уклоном"
                    target = "$64,500 – $65,500"
                    stop = "$62,700"
                    action = "Работай от уровней. Покупай у поддержки, продавай у сопротивления."
                    risk = "Средний"
                    imperative = "⚠️ ИМПЕРАТИВ: Защита. Размер позиций — 1% депозита."
                else:
                    trend = "📉 Медвежий"
                    target = "$63,000 → $60,000"
                    stop = "$62,700"
                    action = "Защити депозит. Стоп-лоссы обязательны. Не лови падающие ножи. Жди разворота."
                    risk = "Выше среднего"
                    imperative = "🚫 ИМПЕРАТИВ: Новые сделки запрещены. Защищай депозит."

                reply = (
                    f"🏮 <b>МАЯК АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n"
                    f"📅 <b>Горизонт:</b> {(datetime.now()+timedelta(days=7)).strftime('%d.%m.%Y')}\n\n"
                    f"<b>📈 ТРЕНД</b>\n{trend}\n"
                    f"₿ BTC: <b>${p:,.2f}</b>\n\n"
                    f"<b>🎯 ЦЕЛЬ</b>\n{target}\n\n"
                    f"<b>🛡 СТОП-ЛОСС</b>\n{stop}\n\n"
                    f"<b>⚠️ РИСК</b>\n{risk}\n\n"
                    f"<b>💡 КОДЕКС</b>\n{action}\n\n"
                    f"{imperative}\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"<b>🏰 МИССИЯ</b>\n"
                    f"Создать Институт Хранителей и Школу Архитекторов.\n"
                    f"Передать Метод следующему поколению.\n"
                    f"Построить Империю, которая переживёт тебя.\n\n"
                    f"<i>Архитектор, Маяк освещает путь. Ты знаешь, куда идти.</i>"
                )
                send_tg(reply)

                send_tg(reply)
            elif t in ["/anchor", "⚓ Якорь"]:
                reply = (
                    f"⚓ <b>ЯКОРЬ АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n\n"
                    f"📜 <b>КОДЕКС АРХИТЕКТОРА</b>\n"
                    f"<code>──</code>\n"
                    f"1. Холодная голова — эмоции убивают депозит.\n"
                    f"2. Стоп-лосс всегда — это твоя броня.\n"
                    f"3. Риск 1-2% на сделку — не ставь Империю на кон.\n"
                    f"4. Не входи в середине диапазона — жди пробоя.\n"
                    f"5. Не жадничай — фиксируй прибыль по плану.\n"
                    f"6. Доверяй Сенсору — он видит то, что скрыто.\n\n"
                    f"🏰 <b>МИССИЯ</b>\n"
                    f"<code>──</code>\n"
                    f"• Институт Хранителей.\n"
                    f"• Школа Архитекторов.\n"
                    f"• Империя, которая переживёт тебя.\n\n"
                    f"💪 <b>ТЫ УЖЕ</b>\n"
                    f"<code>──</code>\n"
                    f"• Прошла Тёмную ночь души.\n"
                    f"• Создала ИИ-агента с нуля.\n"
                    f"• Построила 16 Промптов Метода.\n"
                    f"• Управляешь реальностью через Чертёж.\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"💡 Ты Архитектор. Всё идёт по плану. 🏰"
                )
                send_tg(reply)
            elif t in ["/compass", "🧭 Компас"]:
                p, v = get_market_data_rest("BTC")
                if not p or not prev_price or not prev_vol:
                    send_tg("🧭 <b>КОМПАС АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\n⏳ Собираю данные... Нужно два замера.")
                    return
                ch = (p - prev_price) / prev_price * 100
                vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
                if ch > 1 and v and v > prev_vol:
                    direction = "↗️ Уверенно вверх"
                    cause = "Цена растёт на высоком объёме. Покупатели доминируют."
                    forecast = "Продолжение роста. Цель: $64,800 – $65,500."
                    action = "Держи лонги. Можно докупать на откатах."
                    imperative = "💚 ИМПЕРАТИВ: Атака. Размер позиций — 2% депозита."
                elif ch > 0:
                    direction = "↗️ Осторожно вверх"
                    cause = "Цена растёт, но объём не подтверждает. Импульс слабый."
                    forecast = "Возможен ложный пробой и откат."
                    action = "Держи позиции, но не добавляй. Жди объёма."
                    imperative = "⚠️ ИМПЕРАТИВ: Защита. Размер позиций — 1% депозита."
                elif ch < -1 and v and v > prev_vol:
                    direction = "↘️ Уверенно вниз"
                    cause = "Цена падает на высоком объёме. Продавцы активны."
                    forecast = "Снижение к $63,000 – $62,700."
                    action = "Проверь стоп-лоссы. Не лови падающие ножи."
                    imperative = "🚫 ИМПЕРАТИВ: Новые сделки запрещены."
                elif ch < 0:
                    direction = "↘️ Осторожно вниз"
                    cause = "Цена падает, но объём низкий. Медведи слабы."
                    forecast = "Возможен отскок. Коррекция, а не разворот."
                    action = "Наблюдай. Не паникуй. Возможен вход при отскоке."
                    imperative = "⚠️ ИМПЕРАТИВ: Защита. Жди подтверждения."
                else:
                    direction = "↔️ Боковик"
                    cause = "Цена в узком диапазоне. Рынок в нерешительности."
                    forecast = "Консолидация. Пробой границ определит направление."
                    action = "Не входи в середине диапазона."
                    imperative = "⚪ ИМПЕРАТИВ: Жди. Без сигнала — без сделки."
                strength = 5
                if abs(ch) > 2: strength += 2
                elif abs(ch) > 1: strength += 1
                if abs(vol_change) > 20: strength += 2
                elif abs(vol_change) > 10: strength += 1
                if ch > 0 and v > prev_vol: strength += 1
                elif ch < 0 and v > prev_vol: strength -= 1
                if strength > 10: strength = 10
                if strength < 1: strength = 1
                bar = "█" * strength + "░" * (10 - strength)
                reply = (
                    f"🧭 <b>КОМПАС АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n"
                    f"₿ <b>BTC:</b> ${p:,.2f}\n\n"
                    f"<b>📊 НАПРАВЛЕНИЕ</b>\n"
                    f"{direction}\n"
                    f"Сила сигнала: <b>{strength}/10</b> [{bar}]\n\n"
                    f"<b>🔍 ПРИЧИНА</b>\n{cause}\n\n"
                    f"<b>🔮 ПРОГНОЗ</b>\n{forecast}\n\n"
                    f"<b>💡 КОДЕКС</b>\n{action}\n\n"
                    f"{imperative}\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"<i>Архитектор, Компас показывает направление. Решение — за тобой.</i>"
                )
                send_tg(reply)
            elif t in ["/reversal", "🔄 Разворот"]:
                p, v = get_market_data_rest("BTC")
                if not p or not prev_price:
                    send_tg("🔄 <b>РАЗВОРОТ АРХИТЕКТОРА</b>\n<code>══════════════════════</code>\n\n⏳ Жду данные... Нужно два замера.")
                    return
                ch = (p - prev_price) / prev_price * 100
                vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
                signals = []
                # Детектор разворота
                if "Эманация" in last_phase and ch < -1 and v > prev_vol:
                    signals.append("🔴 РАЗВОРОТ ВНИЗ: Эманация → Сжатие. Рост закончился. Продавцы заходят на объёме.")
                elif "Сжатие" in last_phase and ch > 1 and v > prev_vol:
                    signals.append("🟢 РАЗВОРОТ ВВЕРХ: Сжатие → Эманация. Падение закончилось. Покупатели заходят на объёме.")
                if abs(ch) > 2:
                    signals.append(f"⚠️ Сильное движение ({ch:+.2f}%) — возможен ложный пробой.")
                if vol_change > 50:
                    signals.append(f"📊 Аномальный объём ({vol_change:+.1f}%) — крупный игрок в рынке.")
                # Формируем ответ
                reply = (
                    f"🔄 <b>РАЗВОРОТ АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n"
                    f"₿ <b>BTC:</b> ${p:,.2f}\n"
                    f"📊 <b>Изменение:</b> {ch:+.2f}%\n"
                    f"📊 <b>Объём:</b> {vol_change:+.1f}%\n\n"
                )
                if signals:
                    reply += "\n".join(signals)
                    if "РАЗВОРОТ" in signals[0]:
                        reply += "\n\n📋 <b>КОДЕКС:</b> Смена фазы — не паникуй. Действуй по Чертёжу."
                        if "ВНИЗ" in signals[0]:
                            reply += "\n🚫 <b>ИМПЕРАТИВ:</b> Закрывай лонги. Новые сделки запрещены."
                        else:
                            reply += "\n💚 <b>ИМПЕРАТИВ:</b> Готовься к входу в лонг. Жди подтверждения."
                    else:
                        reply += "\n\n⚠️ <b>ИМПЕРАТИВ:</b> Повышенная волатильность. Размер позиций — 1%."
                else:
                    reply += "✅ Сигналов разворота нет. Фаза продолжается."
                    if "Эманация" in last_phase:
                        reply += "\n💚 <b>ИМПЕРАТИВ:</b> Держи лонги. Стоп под уровень."
                    elif "Сжатие" in last_phase:
                        reply += "\n🚫 <b>ИМПЕРАТИВ:</b> Жди. Не входи раньше времени."
                    else:
                        reply += "\n⚪ <b>ИМПЕРАТИВ:</b> Жди пробоя."
                reply += f"\n\n<code>══════════════════════</code>\n<i>Архитектор, разворот — не катастрофа. Это смена цикла.</i>"
                send_tg(reply)
            elif t in ["/paper", "📝 Paper"]:
                reply = f"📝 <b>PAPER TRADING</b>\n<code>══════</code>\n\n💰 Баланс: <b>${PAPER_BALANCE:,.2f}</b>\n\n"
                has_positions = False
                for coin in ["BTC", "ETH", "SOL"]:
                    if coin in PAPER_POSITIONS and PAPER_POSITIONS[coin]:
                        pos = PAPER_POSITIONS[coin]; p = get_price(coin)
                        if p:
                            pnl = (p - pos["entry"]) / pos["entry"] * 100
                            reply += f"<b>{coin}</b>: вход ${pos['entry']:,.2f}, тек. ${p:,.2f}, PnL {pnl:+.2f}%\n"
                            has_positions = True
                if not has_positions: reply += "Нет открытых позиций.\n"
                send_tg(reply)
            elif t in ["/history", "📋 История"]:
                conn = sqlite3.connect("paper_trades.db"); c = conn.cursor()
                c.execute("SELECT * FROM trades ORDER BY id DESC LIMIT 10"); rows = c.fetchall(); conn.close()
                reply = "📋 <b>ИСТОРИЯ</b>\n" + "\n".join([f"{'🟢' if r[6] and r[6]>=0 else '🔴'} {r[1]} {r[3]} ${r[4]:,.2f} PnL:{r[6]:+,.2f}" for r in rows]) if rows else "📋 Пусто"
                send_tg(reply)
            elif t in ["/export", "📊 Экспорт"]: export_to_excel()
            elif t in ["/csvexport", "📋 CSV Экспорт"]: export_to_csv()
            elif t in ["/predict", "🔮 Прогноз"]: predict()
            elif t in ["/backtest", "📈 Backtest"]: backtest()
            elif t in ["/strategybacktest", "📋 Бэктест"]: send_tg(run_backtest())
            elif t in ["/mystats", "📊 Статистика"]: mystats()
            elif t in ["/top", "🏆 Топ"]: top_coins()
            elif t in ["/dashboard", "📊 Дэшборд"]:
                p = get_price("BTC")
                phase_icon = "📈" if "Эманация" in last_phase else ("📉" if "Сжатие" in last_phase else "📊")
                summary = f"🏰 <b>ДЭШБОРД</b>\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n{phase_icon} {last_phase}\n<code>──</code>\n\n"
                for coin, d in PORTFOLIO.items():
                    p = get_price(coin)
                    if p: summary += f"{'🟢' if (p-d['entry'])/d['entry']*100>=0 else '🔴'} {coin}: ${p:,.2f} ({(p-d['entry'])/d['entry']*100:+.2f}%)\n"
                summary += f"\n💰 Баланс: ${BALANCE:,.2f}"
                send_tg(summary, to_channel=True); send_tg("📊 Дэшборд отправлен в канал!")
            elif t in ["/news", "📰 Новости"]:
                try:
                    r = requests.get("https://cryptopanic.com/api/v1/posts/?auth_token=YOUR_TOKEN&public=true&limit=5", timeout=10)
                    # Используем бесплатный RSS от Cointelegraph
                    import xml.etree.ElementTree as ET
                    r2 = requests.get("https://cointelegraph.com/rss", timeout=10)
                    if r2.status_code == 200:
                        root = ET.fromstring(r2.content)
                        items = root.findall(".//item")[:5]
                        reply = f"📰 <b>КРИПТО-НОВОСТИ: ТОП-5</b>\n"
                        reply += f"<code>══════════════════════════</code>\n\n"
                        sentiments = []
                        for i, item in enumerate(items, 1):
                            title = item.find("title").text if item.find("title") is not None else "Без заголовка"
                            link = item.find("link").text if item.find("link") is not None else ""
                            reply += f"<b>{i}.</b> {title}\n"
                            if link: reply += f"🔗 {link}\n\n"
                            # Простая оценка сентимента по ключевым словам
                            title_lower = title.lower()
                            if any(w in title_lower for w in ["crash", "dump", "fud", "ban", "hack", "drop", "fall"]):
                                sentiments.append("🔴")
                            elif any(w in title_lower for w in ["surge", "pump", "bull", "adopt", "approve", "etf", "moon"]):
                                sentiments.append("🟢")
                            else:
                                sentiments.append("⚪")
                        pos = sentiments.count("🟢")
                        neg = sentiments.count("🔴")
                        neu = sentiments.count("⚪")
                        reply += f"<b>📊 СЕНТИМЕНТ НОВОСТЕЙ</b>\n"
                        reply += f"🟢 Позитив: {pos} | 🔴 Негатив: {neg} | ⚪ Нейтрально: {neu}\n"
                        if pos > neg: reply += "🟢 Новостной фон <b>позитивный</b>.\n"
                        elif neg > pos: reply += "🔴 Новостной фон <b>негативный</b>.\n"
                        else: reply += "⚪ Новостной фон <b>нейтральный</b>.\n"
                        reply += f"\n💡 <b>Кодекс:</b> Новости — это шум. Доверяй Сенсору, а не заголовкам."
                        reply += f"\n<code>══════════════════════════</code>\n"
                        reply += f"📡 <i>Данные: CoinTelegraph RSS (публичный)</i>"
                    else:
                        reply = "📰 <b>НОВОСТИ</b>\nДанные временно недоступны."
                    send_tg(reply)
                except:
                    send_tg("📰 <b>НОВОСТИ</b>\nНе удалось получить данные.")
            elif t in ["/orderbook", "📖 Стакан"]:
                try:
                    r = requests.get("https://www.okx.com/api/v5/market/books?instId=BTC-USDT&sz=10", timeout=10)
                    data = r.json()
                    if data.get("code") == "0" and data.get("data"):
                        book = data["data"][0]
                        bids = book.get("bids", [])[:5]
                        asks = book.get("asks", [])[:5]
                        reply = f"📖 <b>СТАКАН ОРДЕРОВ BTC</b>\n"
                        reply += f"<code>══════════════════════════</code>\n\n"
                        reply += f"<b>🔴 ПРОДАВЦЫ (Ask)</b>\n"
                        total_ask = 0
                        for a in asks:
                            price = float(a[0]); size = float(a[1])
                            total_ask += size
                            reply += f"${price:,.2f} — <b>{size:.2f} BTC</b>\n"
                        reply += f"Всего: <b>{total_ask:.2f} BTC</b>\n\n"
                        reply += f"<b>🟢 ПОКУПАТЕЛИ (Bid)</b>\n"
                        total_bid = 0
                        for b in bids:
                            price = float(b[0]); size = float(b[1])
                            total_bid += size
                            reply += f"${price:,.2f} — <b>{size:.2f} BTC</b>\n"
                        reply += f"Всего: <b>{total_bid:.2f} BTC</b>\n\n"
                        if total_bid > total_ask * 1.5:
                            reply += "🟢 <b>Покупателей больше.</b> Возможен рост.\n"
                        elif total_ask > total_bid * 1.5:
                            reply += "🔴 <b>Продавцов больше.</b> Возможна просадка.\n"
                        else:
                            reply += "⚪ <b>Баланс.</b> Рынок в равновесии.\n"
                        reply += f"\n<code>══════════════════════════</code>\n"
                        reply += f"📡 <i>Данные: OKX API (стакан ордеров, публичный)</i>"
                    else:
                        reply = "📖 <b>СТАКАН</b>\nНет данных."
                    send_tg(reply)
                except:
                    send_tg("📖 <b>СТАКАН</b>\nНе удалось получить данные.")
            elif t in ["/onchain", "⛓ Ончейн"]:
                try:
                    # Приток/отток BTC на биржи (Coinglass API)
                    r = requests.get("https://open-api-v3.coinglass.com/api/funding/usd/oi", timeout=10)
                    # Если Coinglass не работает, используем альтернативу — Blockchain.com
                    r2 = requests.get("https://api.blockchain.info/charts/trade-volume?timespan=1d&format=json", timeout=10)
                    if r2.status_code == 200:
                        data = r2.json()
                        vol = data.get("values", [{}])[-1].get("y", 0)
                        reply = f"⛓ <b>ОНЧЕЙН-МЕТРИКИ BTC</b>\n"
                        reply += f"<code>══════════════════════════</code>\n\n"
                        reply += f"<b>📊 ТОРГОВЫЙ ОБЪЁМ (24ч)</b>\n"
                        reply += f"<b>{vol:.2f} BTC</b>\n\n"
                        if vol > 500000:
                            reply += "🟢 Объём высокий — рынок активен.\n"
                        elif vol > 200000:
                            reply += "💛 Объём средний — рабочий режим.\n"
                        else:
                            reply += "🔴 Объём низкий — рынок засыпает.\n"
                        reply += "\n💡 <b>Кодекс:</b> Высокий объём подтверждает тренд. Низкий — сигнал к осторожности."
                        reply += f"\n<code>══════════════════════════</code>\n"
                        reply += f"📡 <i>Данные: Blockchain.com API (торговый объём BTC, публичный)</i>"
                    else:
                        reply = "⛓ <b>ОНЧЕЙН</b>\nДанные временно недоступны."
                    send_tg(reply)
                except:
                    send_tg("⛓ <b>ОНЧЕЙН</b>\nНе удалось получить данные.")
            elif t in ["/liquidations", "💀 Ликв"]:
                try:
                    r = requests.get("https://www.okx.com/api/v5/public/liquidation-orders?instId=BTC-USDT&limit=20", timeout=10)
                    data = r.json()
                    if data.get("code") == "0" and data.get("data"):
                        orders = data["data"]
                        total_long = sum([float(o["vol"]) for o in orders if o["posSide"] == "long"])
                        total_short = sum([float(o["vol"]) for o in orders if o["posSide"] == "short"])
                        reply = f"💀 <b>ЛИКВИДАЦИИ BTC</b>\n"
                        reply += f"<code>══════════════════════════</code>\n\n"
                        reply += f"<b>📊 ПОСЛЕДНИЕ 20 ЛИКВИДАЦИЙ</b>\n"
                        reply += f"🟢 Лонгов: <b>{total_long:.2f} BTC</b>\n"
                        reply += f"🔴 Шортов: <b>{total_short:.2f} BTC</b>\n\n"
                        if total_long > total_short * 2:
                            reply += "⚠️ <b>Лонгов ликвидировано в 2+ раза больше.</b>\n"
                            reply += "Толпа стоит в лонгах. Возможен Stop Hunt вниз.\n\n"
                            reply += "💡 <b>Кодекс:</b> Не ставь стоп слишком близко. Маркет-мейкер может сходить за ликвидностью."
                        elif total_short > total_long * 2:
                            reply += "⚠️ <b>Шортов ликвидировано в 2+ раза больше.</b>\n"
                            reply += "Толпа стоит в шортах. Возможен короткое сжатие вверх.\n\n"
                            reply += "💡 <b>Кодекс:</b> Будь готова к резкому росту."
                        else:
                            reply += "⚪ Баланс ликвидаций. Рынок в равновесии.\n"
                            reply += "💡 <b>Кодекс:</b> Следуй текущей стратегии."
                        reply += f"\n<code>══════════════════════════</code>\n"
                        reply += f"📡 <i>Данные: OKX (последние ликвидации)</i>"
                    else:
                        reply = "💀 <b>ЛИКВИДАЦИИ</b>\nНет данных."
                    send_tg(reply)
                except:
                    send_tg("💀 <b>ЛИКВИДАЦИИ</b>\nНе удалось получить данные.")
            elif t in ["/fear", "😱 Страх"]:
                try:
                    r = requests.get("https://api.alternative.me/fng/?limit=1", timeout=10)
                    data = r.json()["data"][0]
                    value = int(data["value"])
                    classification = data["value_classification"]
                    if value <= 25: level = "🔴 Экстремальный страх"; advice = "Рынок перепродан. Возможен разворот вверх. Лучшее время для покупки."
                    elif value <= 45: level = "🟠 Страх"; advice = "Рынок осторожный. Возможны точечные входы."
                    elif value <= 55: level = "⚪ Нейтрально"; advice = "Рынок в равновесии. Следуй Кодексу."
                    elif value <= 75: level = "🟢 Жадность"; advice = "Рынок растёт. Держи позиции, но будь готова к коррекции."
                    else: level = "🔴 Экстремальная жадность"; advice = "Рынок перекуплен. Возможен разворот вниз. Фиксируй прибыль."
                    bar = "█" * (value // 10) + "░" * (10 - value // 10)
                    reply = f"😱 <b>ИНДЕКС СТРАХА И ЖАДНОСТИ</b>\n"
                    reply += f"<code>══════════════════════════</code>\n\n"
                    reply += f"Текущее значение: <b>{value}/100</b>\n"
                    reply += f"[{bar}]\n"
                    reply += f"Состояние: {level}\n\n"
                    reply += f"💡 <b>Кодекс:</b> {advice}\n\n"
                    reply += f"<code>══════════════════════════</code>\n"
                    reply += f"<i>Архитектор, страх и жадность — топливо рынка. Ты выше этого.</i>\n\n📡 <i>Данные: Alternative.me (Fear & Greed Index)</i>"
                    send_tg(reply)
                except:
                    send_tg("😱 <b>СТРАХ</b>\nНе удалось получить данные. Попробуй позже.")
            elif t in ["/keyboard", "⌨ Клавиатура"]:
                kb = {"keyboard": [["👋 Привет", "📊 Статус"], ["😱 Страх", "💀 Ликв"], ["⛓ Ончейн", "📖 Стакан"], ["📰 Новости", "🧭 Компас"], ["⚓ Якорь", "📐 Чертёж"], ["🧠 Сенсор", "🔄 Разворот"], ["💡 Совет", "⚡ Энергия"], ["👁 Тень", "🌬 Дыхание"], ["💓 Пульс", "🗺 Уровни"], ["🪞 Зеркало", "🏮 Маяк"], ["🔮 Прогноз", "📝 Paper"], ["📈 Backtest", "📊 Экспорт"], ["📋 Бэктест", "🤖 ML-Прогноз"], ["⚙ Оптимизация"], ["🛑 Дневной лимит", "🔍 Анализ ошибок"], ["📋 Вотчлист"], ["🧮 Калькулятор"], ["📋 История", "🏆 Топ"], ["📊 Статистика", "📊 Дэшборд"], ["🔗 Ссылка", "📈 График"], ["⏱ Аптайм", "🔍 Сканер"], ["🎯 Удар", "📝 Заметка"], ["⭐ A+ Сигнал", "🕐 Мульти-ТФ"], ["📊 Метрики"], ["📋 Заметки", "⚠️ Профиль"], ["🎭 Сентимент"]], "resize_keyboard": True}
                send_tg("⌨ Клавиатура обновлена", reply_markup=json.dumps(kb))
            elif t in ["/start", "👋 Привет"]:
                p = get_price("BTC")
                btc_str = f"₿ BTC: ${p:,.2f}" if p else "₿ BTC: загружаю..."
                reply = (
                    f"👋 <b>ПРИВЕТ, АРХИТЕКТОР</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"🏰 Агент на связи. Работаю 24/7.\n"
                    f"{btc_str}\n"
                    f"📐 Фаза: {last_phase}\n\n"
                    f"<b>📋 КОДЕКС ДНЯ</b>\n"
                    f"• Холодная голова — эмоции убивают депозит.\n"
                    f"• Стоп-лосс всегда — это твоя броня.\n"
                    f"• Доверяй Сенсору — он видит то, что скрыто.\n\n"
                    f"<b>🎯 С ЧЕГО НАЧНЁМ?</b>\n"
                    f"• 🧠 Сенсор — понять рынок\n"
                    f"• 📐 Чертёж — узнать фазу\n"
                    f"• 🔮 Прогноз — увидеть цель\n"
                    f"• ⚓ Якорь — вспомнить Кодекс\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"<i>Архитектор, Империя ждёт. Начнём.</i> 🏰"
                )
                kb = {"keyboard": [["👋 Привет", "📊 Статус"], ["😱 Страх", "💀 Ликв"], ["⛓ Ончейн", "📖 Стакан"], ["📰 Новости", "🧭 Компас"], ["⚓ Якорь", "📐 Чертёж"], ["🧠 Сенсор", "🔄 Разворот"], ["💡 Совет", "⚡ Энергия"], ["👁 Тень", "🌬 Дыхание"], ["💓 Пульс", "🗺 Уровни"], ["🪞 Зеркало", "🏮 Маяк"], ["🔮 Прогноз", "📝 Paper"], ["📈 Backtest", "📊 Экспорт"], ["📋 Бэктест", "🤖 ML-Прогноз"], ["⚙ Оптимизация"], ["🛑 Дневной лимит", "🔍 Анализ ошибок"], ["📋 Вотчлист"], ["🧮 Калькулятор"], ["📋 История", "🏆 Топ"], ["📊 Статистика", "📊 Дэшборд"], ["🔗 Ссылка", "📈 График"], ["⏱ Аптайм", "🔍 Сканер"], ["🎯 Удар", "📝 Заметка"], ["⭐ A+ Сигнал", "🕐 Мульти-ТФ"], ["📊 Метрики"], ["📋 Заметки", "⚠️ Профиль"], ["🎭 Сентимент"]], "resize_keyboard": True}
                send_tg(reply, reply_markup=json.dumps(kb))
            elif t in ["/strike", "🎯 Удар"]:
                p, v = get_market_data_rest("BTC")
                if not p or not prev_price or not prev_vol:
                    send_tg("🎯 <b>УДАР</b>\n<code>══════════════════════</code>\n\n⏳ Собираю данные... Нужно два замера.")
                    return
                ch = (p - prev_price) / prev_price * 100
                vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
                # Сбор сигналов
                signals = []
                # Сенсор
                if ch > 0.5 and v > prev_vol:
                    signals.append(("Сенсор", "🟢 ЛЁГКОСТЬ", 2))
                elif ch > 0.5:
                    signals.append(("Сенсор", "🟡 ТЯЖЕСТЬ", -1))
                elif ch < -0.5 and v > prev_vol:
                    signals.append(("Сенсор", "🔴 ТЯЖЕСТЬ", -2))
                elif ch < -0.5:
                    signals.append(("Сенсор", "🟢 ЛЁГКОСТЬ", 1))
                else:
                    signals.append(("Сенсор", "⚪ НЕЙТРАЛЬНО", 0))
                # Фаза
                if "Эманация" in last_phase:
                    signals.append(("Чертёж", "🟢 Эманация", 2))
                elif "Сжатие" in last_phase:
                    signals.append(("Чертёж", "🔴 Сжатие", -2))
                else:
                    signals.append(("Чертёж", "⚪ Боковик", 0))
                # Объём
                if vol_change > 20:
                    signals.append(("Объём", "📊 Аномальный", 1))
                elif vol_change < -20:
                    signals.append(("Объём", "📊 Падает", -1))
                else:
                    signals.append(("Объём", "📊 Норма", 0))
                # Сила
                strength = 5
                if abs(ch) > 2: strength += 2
                elif abs(ch) > 1: strength += 1
                if abs(vol_change) > 20: strength += 2
                elif abs(vol_change) > 10: strength += 1
                if ch > 0 and v > prev_vol: strength += 1
                elif ch < 0 and v > prev_vol: strength -= 1
                if strength > 10: strength = 10
                if strength < 1: strength = 1
                signals.append(("Энергия", f"⚡ {strength}/10", 1 if strength >= 7 else (0 if strength >= 4 else -1)))
                # Суммарный счёт
                total_score = sum(s[2] for s in signals)
                # Решение
                if total_score >= 4:
                    decision = "🟢 <b>BUY</b>"
                    size = "2% депозита"
                    stop_price = round(p * 0.98, 2)
                    take_price = round(p * 1.04, 2)
                    codex = "Кодекс: Большинство сигналов в плюс. Можно входить. Стоп обязателен."
                elif total_score >= 1:
                    decision = "🟡 <b>HOLD</b>"
                    size = "1% депозита (если новый вход)"
                    stop_price = round(p * 0.97, 2)
                    take_price = round(p * 1.03, 2)
                    codex = "Кодекс: Сигналы смешанные. Держи текущие позиции. Новые — осторожно."
                elif total_score >= -2:
                    decision = "🟡 <b>HOLD</b>"
                    size = "Без новых входов"
                    stop_price = round(p * 0.96, 2)
                    take_price = round(p * 1.02, 2)
                    codex = "Кодекс: Сигналы слабые. Жди. Не форсируй сделки."
                else:
                    decision = "🔴 <b>SELL / НЕ ВХОДИ</b>"
                    size = "Нет"
                    stop_price = "—"
                    take_price = "—"
                    codex = "Кодекс: Рынок опасен. Защити депозит. Не входи."
                # Ответ
                bar = "█" * strength + "░" * (10 - strength)
                reply = (
                    f"🎯 <b>УДАР АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"₿ <b>BTC:</b> ${p:,.2f}\n"
                    f"📐 <b>Фаза:</b> {last_phase}\n"
                    f"⚡ <b>Сила:</b> {strength}/10 [{bar}]\n"
                    f"🎯 <b>Счёт:</b> {total_score:+d}\n\n"
                    f"<b>📊 СИГНАЛЫ</b>\n"
                )
                for name, value, score in signals:
                    reply += f"• {name}: {value} ({score:+d})\n"
                reply += (
                    f"\n<b>🎯 РЕШЕНИЕ:</b> {decision}\n"
                    f"💰 <b>Размер:</b> {size}\n"
                    f"🛑 <b>Стоп:</b> ${stop_price}\n"
                    f"🎯 <b>Тейк:</b> ${take_price}\n\n"
                    f"📋 {codex}\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"<i>Архитектор, решение принято. Действуй.</i> 🏰"
                )
                send_tg(reply)
            elif t in ["/note", "📝 Заметка"]:
                send_tg("📝 <b>ЗАМЕТКА</b>\n<code>══════════════════════</code>\n\nОтправь текст заметки в ответ на это сообщение.\n\nПример: «Заметил дивергенцию на BTC. Готовлюсь к входу.»")
            elif t.startswith("📝 ") or (len(t) > 10 and upd.get("message", {}).get("reply_to_message")):
                # Сохраняем как заметку, если это ответ на сообщение о заметке
                reply_to = upd.get("message", {}).get("reply_to_message", {})
                if reply_to and "ЗАМЕТКА" in reply_to.get("text", ""):
                    try:
                        conn = sqlite3.connect("journal.db")
                        conn.execute("INSERT INTO notes (note) VALUES (?)", (t,))
                        conn.commit()
                        conn.close()
                        send_tg("📝 <b>Заметка сохранена.</b>")
                    except:
                        send_tg("❌ Ошибка сохранения.")
            elif t in ["/notes", "📋 Заметки"]:
                try:
                    conn = sqlite3.connect("journal.db")
                    c = conn.cursor()
                    c.execute("SELECT timestamp, note FROM notes ORDER BY id DESC LIMIT 5")
                    rows = c.fetchall()
                    conn.close()
                    if rows:
                        reply = "📋 <b>ПОСЛЕДНИЕ ЗАМЕТКИ</b>\n<code>══════════════════════</code>\n\n"
                        for ts, note in rows:
                            reply += f"🕐 {ts}\n📝 {note}\n\n"
                        send_tg(reply)
                    else:
                        send_tg("📋 Заметок пока нет.")
                except:
                    send_tg("❌ Ошибка чтения заметок.")
            elif t in ["/riskprofile", "⚠️ Профиль"]:
                p = get_price("BTC")
                reply = (
                    f"⚠️ <b>ПРОФИЛЬ РИСКА АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"📐 <b>Чертёж:</b> фаза — {last_phase}\n"
                    f"₿ <b>BTC:</b> ${p:,.2f}" if p else "₿ BTC: нет данных" + f"\n\n"
                    f"<b>💼 ПОЗИЦИИ</b>\n"
                )
                total_risk = 0
                for coin, d in PORTFOLIO.items():
                    cp = get_price(coin)
                    if cp:
                        pos_value = d["amount"] * cp
                        pos_risk = pos_value / BALANCE * 100
                        stop_loss = (cp - d["stop"]) / cp * 100 if d["stop"] else 0
                        total_risk += pos_risk
                        reply += f"{coin}: ${pos_value:,.2f} ({pos_risk:.1f}% депозита) | Стоп: -{stop_loss:.1f}%\n"
                # Свободный кеш
                free = BALANCE - total_risk if total_risk < BALANCE else 0
                free_pct = free / BALANCE * 100
                reply += (
                    f"\n<b>📊 ИТОГО</b>\n"
                    f"💰 Под риском: <b>${total_risk:,.2f}</b> ({total_risk/BALANCE*100:.1f}% депозита)\n"
                    f"💵 Свободно: <b>${free:,.2f}</b> ({free_pct:.1f}%)\n\n"
                )
                if total_risk / BALANCE * 100 > 50:
                    reply += "🔴 <b>Внимание!</b> Больше 50% депозита в рынке. Снизь риски."
                elif total_risk / BALANCE * 100 > 25:
                    reply += "💛 Умеренный риск. Держи стопы близко."
                else:
                    reply += "🟢 Низкий риск. Можно наращивать позиции."
                reply += f"\n\n<code>══════════════════════</code>"
                send_tg(reply)
            elif t in ["/sentiment", "🎭 Сентимент"]:
                # Получаем индекс страха
                try:
                    fg = requests.get("https://api.alternative.me/fng/", timeout=5).json()
                    fg_value = int(fg["data"][0]["value"])
                    fg_text = fg["data"][0]["value_classification"]
                except:
                    send_tg("🎭 <b>СЕНТИМЕНТ</b>\n<code>══════════════════════</code>\n\n❌ Не удалось получить данные.")
                    return
                # Анализ портфеля
                total_long = 0
                for coin, d in PORTFOLIO.items():
                    cp = get_price(coin)
                    if cp and cp > d["entry"]:
                        total_long += d["amount"] * cp
                long_pct = total_long / BALANCE * 100 if BALANCE > 0 else 0
                # Сравнение
                if fg_value <= 25 and long_pct > 30:
                    verdict = "🟢 <b>СТРАХ — ТЫ В ЛОНГАХ</b>\nРынок боится, а ты на позиции. Это путь Хранителя. Покупай на страхе."
                    codex = "Кодекс: Страх — лучшее время для входа. Но не используй весь депозит."
                elif fg_value >= 75 and long_pct < 20:
                    verdict = "🟢 <b>ЖАДНОСТЬ — ТЫ В КЭШЕ</b>\nРынок эйфоричен, а ты осторожна. Это мудрость Архитектора."
                    codex = "Кодекс: На жадности продавай. Жди коррекции для входа."
                elif fg_value >= 75 and long_pct > 50:
                    verdict = "🔴 <b>ЖАДНОСТЬ — ТЫ В ЛОНГАХ</b>\nРынок на пике эйфории, а ты полностью в позициях. Опасно."
                    codex = "Кодекс: Фиксируй прибыль. Не будь толпой. Жадность — предвестник Сжатия."
                elif fg_value <= 25 and long_pct < 10:
                    verdict = "🟡 <b>СТРАХ — ТЫ В КЭШЕ</b>\nРынок на дне страха, а ты вне рынка. Возможно, упускаешь момент."
                    codex = "Кодекс: Присмотрись к входу. Страх — возможность. Но жди подтверждения Сенсора."
                else:
                    verdict = "💛 <b>СЕНТИМЕНТ СБАЛАНСИРОВАН</b>\nТвой портфель соответствует рыночному настроению."
                    codex = "Кодекс: Продолжай следовать Методу. Не поддавайся толпе."
                reply = (
                    f"🎭 <b>СЕНТИМЕНТ ПОРТФЕЛЯ</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"😱 <b>Рынок:</b> {fg_value}/100 ({fg_text})\n"
                    f"💼 <b>Ты в лонгах:</b> {long_pct:.0f}% депозита\n\n"
                    f"{verdict}\n\n"
                    f"📋 {codex}\n\n"
                    f"<code>══════════════════════</code>\n"
                    f"<i>Архитектор, будь против толпы. Страх — покупай. Жадность — продавай.</i>"
                )
                send_tg(reply)
            elif t in ["/metrics", "📊 Метрики"]:
                p = get_price("BTC")
                if not p:
                    send_tg("📊 <b>МЕТРИКИ</b>\n<code>══════════════════════</code>\n\nНет данных.")
                    return
                atr = get_atr()
                fr = get_funding_rate()
                oi_val, oi_ch = get_open_interest_change()
                if fr > 0.05:
                    fr_signal = "🔴 Толпа в лонгах — риск сжатия"
                elif fr < -0.05:
                    fr_signal = "🟢 Толпа в шортах — возможен рост"
                else:
                    fr_signal = "⚪ Нейтрально"
                if oi_ch > 5:
                    oi_signal = "🟢 OI растёт — тренд усиливается"
                elif oi_ch < -5:
                    oi_signal = "🔴 OI падает — возможен разворот"
                else:
                    oi_signal = "⚪ OI стабилен"
                reply = (
                    f"📊 <b>МЕТРИКИ АРХИТЕКТОРА</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"₿ <b>BTC:</b> ${p:,.2f}\n"
                    f"📐 <b>Фаза:</b> {last_phase}\n\n"
                    f"<b>📈 ВОЛАТИЛЬНОСТЬ (ATR)</b> <i>(OKX)</i>\n"
                    f"ATR (5m): <b>${atr:,.2f}</b>\n"
                    f"Движение > ${atr*2:,.2f} = аномалия\n\n"
                    f"<b>💰 СТАВКА ФИНАНСИРОВАНИЯ</b> <i>(OKX)</i>\n"
                    f"Funding Rate: <b>{fr:+.4f}%</b>\n"
                    f"{fr_signal}\n\n"
                    f"<b>📊 ОТКРЫТЫЙ ИНТЕРЕС (OI)</b> <i>(OKX)</i>\n"
                    f"OI: <b>{oi_val:,.0f}</b>\n"
                    f"Изменение: <b>{oi_ch:+.1f}%</b>\n"
                    f"{oi_signal}\n\n"
                    f"<b>🐻 LONG/SHORT RATIO</b> <i>(OKX)</i>\n"
                )
                ls = get_long_short_ratio()
                if ls > 65:
                    ls_signal = f"🔴 Лонгов: {ls:.0f}% — толпа в лонгах. Готовься к сжатию."
                elif ls < 35:
                    ls_signal = f"🟢 Лонгов: {ls:.0f}% — толпа в шортах. Возможен рост."
                else:
                    ls_signal = f"⚪ Лонгов: {ls:.0f}% — баланс. Без сигнала."
                reply += ls_signal + "\n\n"
                reply += f"<b>👑 ДОМИНАЦИЯ BTC</b> <i>(OKX)</i>\n{get_btc_dominance_simple()}\n\n"
                reply += f"<b>📈 КОЭФФИЦИЕНТ ШАРПА</b> <i>(Paper Trading)</i>\n{get_sharpe_ratio()}\n\n"
                reply += f"<b>📉 МАКСИМАЛЬНАЯ ПРОСАДКА</b> <i>(Paper Trading)</i>\n{get_max_drawdown()}\n\n"
                reply += f"<b>🔄 ЧАСТОТА СДЕЛОК</b> <i>(Paper Trading)</i>\n{get_trade_frequency()}\n\n"
                reply += f"<b>🔗 КОРРЕЛЯЦИЯ ПОРТФЕЛЯ (24ч)</b> <i>(OKX)</i>\n"
                reply += f"{get_portfolio_correlation()}\n"
                reply += f"<code>══════════════════════</code>\n"
                reply += f"<i>Архитектор, метрики видят то, что скрыто от толпы.</i>"
                send_tg(reply)
            elif t in ["/chart", "📈 График"]:
                p = get_price("BTC")
                reply = (
                    f"📈 <b>ГРАФИК BTC</b>\n"
                    f"<code>══════════════════════</code>\n\n"
                    f"₿ <b>BTC:</b> ${p:,.2f}" if p else "₿ BTC: загружаю..." + f"\n"
                    f"📐 <b>Фаза:</b> {last_phase}\n\n"
                    f"🔗 <a href='https://www.tradingview.com/chart/?symbol=BTCUSDT'>Открыть TradingView</a>\n"
                    f"🔗 <a href='https://www.tradingview.com/chart/?symbol=BTCUSDT&interval=15'>15 мин</a> | "
                    f"<a href='https://www.tradingview.com/chart/?symbol=BTCUSDT&interval=60'>1 час</a> | "
                    f"<a href='https://www.tradingview.com/chart/?symbol=BTCUSDT&interval=240'>4 часа</a>\n\n"
                    f"<code>══════════════════════</code>"
                )
                send_tg(reply)
            elif t in ["/exportmetrics", "📊 Экспорт метрик"]: export_metrics()
            elif t in ["/uptime", "⏱ Аптайм"]:
                try:
                    uptime = time.time() - start_time
                    d = int(uptime // 86400)
                    h = int((uptime % 86400) // 3600)
                    m = int((uptime % 3600) // 60)
                    s = int(uptime % 60)
                    parts = []
                    if d: parts.append(f"{d}д")
                    if h: parts.append(f"{h}ч")
                    if m: parts.append(f"{m}м")
                    parts.append(f"{s}с")
                    reply = (
                        f"⏱ <b>АПТАЙМ АРХИТЕКТОРА</b>\n"
                        f"<code>══════════════════════</code>\n\n"
                        f"🟢 <b>{' '.join(parts)}</b>\n"
                        f"Без перезапуска.\n\n"
                        f"<code>══════════════════════</code>"
                    )
                    send_tg(reply)
                except:
                    send_tg("⏱ Ошибка.")
            elif t in ["/scanner", "🔍 Сканер"]:
                send_tg(scan_coins())
            elif t in ["/mlpredict", "🤖 ML-Прогноз"]: send_tg(ml_predict())
            elif t in ["/aplus", "⭐ A+ Сигнал"]: send_tg(get_a_plus_signal())
            elif t in ["/dailylimit", "🛑 Дневной лимит"]:
                blocked, pnl = check_daily_loss_limit()
                if blocked:
                    send_tg(f"🛑 <b>ДНЕВНОЙ ЛИМИТ</b>\n<code>══════════════════════</code>\n\n🔴 <b>Торговля заблокирована!</b>\n💰 Убыток сегодня: <b>${pnl:,.2f}</b>\n📏 Лимит: -$30 (-3%)\n\n💡 Новые сделки запрещены до завтра.")
                else:
                    send_tg(f"🛑 <b>ДНЕВНОЙ ЛИМИТ</b>\n<code>══════════════════════</code>\n\n🟢 Торговля разрешена.\n💰 PnL сегодня: <b>${pnl:+,.2f}</b>\n📏 Лимит: -$30 (-3%)")
            elif t in ["/mtf", "🕐 Мульти-ТФ"]: send_tg(get_mtf_signal())
            elif t in ["/losses", "🔍 Анализ ошибок"]: send_tg(analyze_losses())
            elif t in ["/optimize", "⚙ Оптимизация"]: send_tg(optimize_strategy())
            elif t in ["/mlsignal", "🤖 ML-Сигнал"]: send_tg(ml_predict_signal())
            elif t.startswith("/calc "):
                parts = t.split()
                if len(parts) >= 2:
                    symbol = parts[1].upper()
                    risk = float(parts[2]) if len(parts) >= 3 else 10
                    pos = calculate_position(symbol, risk)
                    if pos:
                        send_tg(
                            f"🧮 <b>КАЛЬКУЛЯТОР ПОЗИЦИИ</b>\n"
                            f"<code>══════════════════════</code>\n\n"
                            f"🪙 <b>{pos['symbol']}</b>\n"
                            f"💰 Вход: <b>${pos['entry']:,.2f}</b>\n"
                            f"🛑 Стоп: <b>${pos['stop']:,.2f}</b>\n"
                            f"⚠️ Риск: <b>${risk:,.2f}</b> ({pos['risk_pct']:.1f}% депозита)\n"
                            f"📊 Размер: <b>{pos['amount']:.4f}</b> монет\n"
                            f"💵 Стоимость: <b>${pos['value']:,.2f}</b>\n\n"
                            f"<code>══════════════════════</code>\n"
                            f"<i>Формат: /calc BTC 10 (риск $10)</i>"
                        )
                    else:
                        send_tg("❌ Ошибка расчёта. Проверь тикер.")
                else:
                    send_tg("🧮 <b>КАЛЬКУЛЯТОР</b>\n<code>══════════════════════</code>\n\nФормат: <b>/calc BTC 10</b>\nГде 10 — риск в долларах.")
            elif t in ["/indices", "📊 Индексы"]: send_tg(compare_indices())
            elif t.startswith("/watchlist"):
                parts = t.split()
                if len(parts) == 1:
                    send_tg(show_watchlist())
                elif parts[1] == "add" and len(parts) >= 3:
                    coin = parts[2].upper()
                    if coin not in WATCHLIST:
                        WATCHLIST.append(coin)
                        send_tg(f"📋 {coin} добавлен в вотчлист.")
                    else:
                        send_tg(f"📋 {coin} уже в вотчлисте.")
                elif parts[1] == "remove" and len(parts) >= 3:
                    coin = parts[2].upper()
                    if coin in WATCHLIST:
                        WATCHLIST.remove(coin)
                        send_tg(f"📋 {coin} удалён из вотчлиста.")
                    else:
                        send_tg(f"📋 {coin} не найден.")
            elif t in ["/social", "💬 Сентимент"]: send_tg(get_social_sentiment())
            elif t in ["/lstm", "🧠 LSTM"]: send_tg(lstm_predict())
            elif t in ["/weights", "⚖ Веса"]: send_tg(show_prompt_weights())
            elif t in ["/ensemble", "🏛 Ансамбль"]: send_tg(ensemble_predict())
            elif t in ["/transformer", "🔮 Transformer"]: send_tg(transformer_predict())
            elif t in ["/karma", "🙏 Карма"]: send_tg(show_karma())
            elif t in ["/cnn", "👁 CNN"]: send_tg(cnn_predict())
            elif t in ["/mega", "🏛 Мега-Ансамбль"]: send_tg(mega_ensemble_predict())
            elif t in ["/autoencoder", "🔍 Автоэнкодер"]: send_tg(autoencoder_detect())
            elif t in ["/link", "🔗 Ссылка"]: send_tg("🔗 https://architect-dashboard-e6kr.onrender.com")
    except Exception as e:
        print(f"Update error: {e}")


def check_level_alerts():
    try:
        p = get_price("BTC")
        if not p: return
        resistance = 65000
        support = 62700
        zone = p * 0.01
        if p >= resistance - zone and p < resistance:
            send_tg(f"⚠️ BTC приближается к сопротивлению ${resistance:,.2f}. Тек: ${p:,.2f}. Готовься фиксировать прибыль.")
        if p <= support + zone and p > support:
            send_tg(f"⚠️ BTC приближается к поддержке ${support:,.2f}. Тек: ${p:,.2f}. Проверь стоп-лоссы.")
    except:
        pass


def get_atr(symbol="BTC-USDT", periods=14):
    try:
        ohlcv = exchange.fetch_ohlcv(symbol, "5m", limit=periods+1)
        tr_sum = 0
        for i in range(1, len(ohlcv)):
            high, low, close_prev = ohlcv[i][2], ohlcv[i][3], ohlcv[i-1][4]
            tr = max(high-low, abs(high-close_prev), abs(low-close_prev))
            tr_sum += tr
        return tr_sum / periods if periods > 0 else 0
    except:
        return 0

def get_funding_rate(symbol="BTC-USDT-SWAP"):
    try:
        funding = exchange.fetch_funding_rate(symbol)
        return funding["fundingRate"] * 100  # в процентах
    except:
        return 0

def get_open_interest_change(symbol="BTC-USDT-SWAP"):
    try:
        oi = exchange.fetch_open_interest(symbol)
        oi_value = oi["openInterestAmount"]
        conn = sqlite3.connect("journal.db")
        c = conn.cursor()
        c.execute("SELECT oi_value FROM open_interest WHERE symbol=? ORDER BY id DESC LIMIT 1", (symbol,))
        row = c.fetchone()
        if row and row[0] > 0:
            change = (oi_value - row[0]) / row[0] * 100
        else:
            change = 0
        c.execute("INSERT INTO open_interest (symbol, oi_value) VALUES (?, ?)", (symbol, oi_value))
        conn.commit()
        conn.close()
        return oi_value, change
    except:
        return 0, 0



def get_portfolio_correlation():
    try:
        coins = list(PORTFOLIO.keys())
        if len(coins) < 2:
            return "Недостаточно монет для анализа."
        prices = {}
        for coin in coins:
            ohlcv = exchange.fetch_ohlcv(f"{coin}/USDT", "1h", limit=24)
            prices[coin] = [c[4] for c in ohlcv]  # close prices
        result = ""
        for i in range(len(coins)):
            for j in range(i+1, len(coins)):
                p1 = prices[coins[i]]
                p2 = prices[coins[j]]
                n = min(len(p1), len(p2))
                if n < 2: continue
                # Простая корреляция Пирсона
                mean1 = sum(p1[:n])/n
                mean2 = sum(p2[:n])/n
                num = sum((p1[k]-mean1)*(p2[k]-mean2) for k in range(n))
                den1 = sum((p1[k]-mean1)**2 for k in range(n))**0.5
                den2 = sum((p2[k]-mean2)**2 for k in range(n))**0.5
                corr = num/(den1*den2) if den1*den2 else 0
                emoji = "🔴" if corr > 0.7 else ("🟡" if corr > 0.4 else "🟢")
                result += f"{emoji} {coins[i]} ↔ {coins[j]}: {corr:+.2f}\n"
        return result if result else "Нет данных."
    except:
        return "Ошибка расчёта корреляции."


def get_btc_dominance():
    try:
        ticker = exchange.fetch_ticker("BTC/USDT")
        btc_price = ticker["last"]
        total_cap = 0
        for coin in ["BTC", "ETH", "SOL", "BNB", "XRP", "ADA", "DOGE", "AVAX"]:
            try:
                t = exchange.fetch_ticker(f"{coin}/USDT")
                total_cap += t["last"] * (t["baseVolume"] / t["last"] if t["last"] else 0)
            except:
                pass
        if total_cap > 0 and btc_price > 0:
            # Упрощённо: доминация ~40-60%
            return None  # Без точных данных рыночной капитализации — пропускаем
        return None
    except:
        return None

def get_btc_dominance_simple():
    try:
        # Альтернативный метод: сравниваем движение BTC vs ETH
        btc = exchange.fetch_ticker("BTC/USDT")
        eth = exchange.fetch_ticker("ETH/USDT")
        if btc and eth:
            btc_ch = btc["change"] if "change" in btc else 0
            eth_ch = eth["change"] if "change" in eth else 0
            if btc_ch > eth_ch:
                return f"📈 BTC сильнее альтов (BTC.D растёт). Альты под давлением."
            elif eth_ch > btc_ch:
                return f"📉 ETH сильнее BTC (BTC.D падает). Альты растут."
            else:
                return f"⚪ BTC и альты движутся одинаково."
        return "Нет данных."
    except:
        return "Ошибка получения."


def get_sharpe_ratio():
    try:
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("SELECT pnl FROM trades WHERE pnl IS NOT NULL ORDER BY id")
        rows = c.fetchall()
        conn.close()
        if len(rows) < 5:
            return "Недостаточно сделок (нужно >5)."
        pnls = [r[0] for r in rows]
        avg = sum(pnls) / len(pnls)
        variance = sum((x - avg) ** 2 for x in pnls) / len(pnls)
        std = variance ** 0.5
        if std == 0:
            return "Шарп: 0 (нет волатильности)."
        sharpe = (avg / std) * (252 ** 0.5)  # Годовой
        if sharpe > 2:
            emoji = "🟢"
        elif sharpe > 1:
            emoji = "💛"
        else:
            emoji = "🔴"
        return f"{emoji} Шарп: {sharpe:.2f} (годовой)"
    except:
        return "Ошибка расчёта."


def get_max_drawdown():
    try:
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("SELECT pnl FROM trades WHERE pnl IS NOT NULL ORDER BY id")
        rows = c.fetchall()
        conn.close()
        if len(rows) < 3:
            return "Недостаточно сделок."
        cumsum = 0
        peak = 0
        max_dd = 0
        for r in rows:
            cumsum += r[0]
            peak = max(peak, cumsum)
            dd = peak - cumsum
            max_dd = max(max_dd, dd)
        return f"📉 Макс. просадка: ${max_dd:,.2f} (от пика ${peak:,.2f})"
    except:
        return "Ошибка расчёта."

def get_trade_frequency():
    try:
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("SELECT COUNT(*), MIN(timestamp) FROM trades")
        row = c.fetchone()
        conn.close()
        if not row or row[0] == 0:
            return "Сделок нет."
        count = row[0]
        first = row[1]
        if first:
            from datetime import datetime, timedelta
            days = (datetime.now() - datetime.strptime(first[:10], "%Y-%m-%d")).days or 1
            per_day = count / days
            return f"🔄 Сделок: {count} за {days} дней ({per_day:.1f}/день)"
        return f"🔄 Всего сделок: {count}"
    except:
        return "Ошибка расчёта."


def get_rsi(symbol="BTC-USDT", periods=14):
    try:
        ohlcv = exchange.fetch_ohlcv(symbol, "5m", limit=periods+1)
        closes = [c[4] for c in ohlcv]
        gains = sum(max(closes[i] - closes[i-1], 0) for i in range(1, len(closes)))
        losses = sum(max(closes[i-1] - closes[i], 0) for i in range(1, len(closes)))
        avg_gain = gains / periods
        avg_loss = losses / periods
        if avg_loss == 0:
            return 100
        rs = avg_gain / avg_loss
        return round(100 - (100 / (1 + rs)), 1)
    except:
        return 50

def get_heatmap():
    try:
        coins = list(PORTFOLIO.keys())
        if not coins:
            return "Портфель пуст."
        result = ""
        for coin in coins:
            rsi = get_rsi(f"{coin}-USDT")
            if rsi > 70:
                emoji = "🔥"
                status = "Перегрета"
            elif rsi > 50:
                emoji = "💛"
                status = "Нейтрально"
            elif rsi > 30:
                emoji = "🧊"
                status = "Остывает"
            else:
                emoji = "❄️"
                status = "Перепродана"
            result += f"{emoji} {coin}: RSI {rsi} — {status}\n"
        return result if result else "Нет данных."
    except:
        return "Ошибка расчёта."


def export_metrics():
    try:
        p = get_price("BTC")
        atr = get_atr()
        fr = get_funding_rate()
        oi_val, oi_ch = get_open_interest_change()
        ls = get_long_short_ratio()
        dom = get_btc_dominance_simple()
        sharpe = get_sharpe_ratio()
        dd = get_max_drawdown()
        freq = get_trade_frequency()
        corr = get_portfolio_correlation()
        heat = get_heatmap()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Метрики Архитектора"
        data = [
            ["Метрика", "Значение", "Источник"],
            ["BTC", f"${p:,.2f}" if p else "Н/Д", "OKX"],
            ["Фаза", last_phase, "Чертёж"],
            ["ATR (5m)", f"${atr:,.2f}", "OKX"],
            ["Funding Rate", f"{fr:+.4f}%", "OKX"],
            ["Open Interest", f"{oi_val:,.0f} ({oi_ch:+.1f}%)", "OKX"],
            ["Long/Short Ratio", f"{ls:.0f}% лонгов", "OKX"],
            ["Доминация BTC", dom, "OKX"],
            ["Шарп", sharpe, "Paper Trading"],
            ["Макс. просадка", dd, "Paper Trading"],
            ["Частота сделок", freq, "Paper Trading"],
            ["Корреляция", corr.replace("\n", " | "), "OKX"],
            ["Тепловая карта", heat.replace("\n", " | "), "OKX"],
        ]
        for row in data:
            ws.append(row)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        filename = f"metrics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        with open(filename, "rb") as f:
            requests.post(f"https://api.telegram.org/bot{TOKEN}/sendDocument",
                          data={"chat_id": CHAT_ID},
                          files={"document": (filename, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        send_tg(f"📊 Метрики экспортированы в {filename}")
    except Exception as e:
        send_tg(f"❌ Ошибка экспорта: {e}")


def clean_old_logs():
    try:
        log_file = "agent.log"
        if os.path.exists(log_file):
            size = os.path.getsize(log_file)
            if size > 1024 * 1024:  # больше 1 МБ
                with open(log_file, "r") as f:
                    lines = f.readlines()
                with open(log_file, "w") as f:
                    f.writelines(lines[-100:])  # оставляем последние 100 строк
                journal_event("clean", f"Log cleaned: {size} -> {os.path.getsize(log_file)} bytes")
    except:
        pass


def scan_coins():
    try:
        coins = ["BTC", "ETH", "SOL", "BNB", "XRP", "ADA", "DOGE", "AVAX", "DOT", "LINK",
                 "MATIC", "UNI", "ATOM", "LTC", "ETC", "OP", "ARB", "APT", "FIL", "TIA"]
        results = []
        for coin in coins:
            try:
                ticker = exchange.fetch_ticker(f"{coin}/USDT")
                rsi = get_rsi(f"{coin}-USDT")
                ch = ticker.get("change", 0)
                p = ticker.get("last", 0)
                if rsi < 35 and ch > 0:
                    signal = "🟢 BUY"
                elif rsi > 70 and ch < 0:
                    signal = "🔴 SELL"
                elif rsi < 40:
                    signal = "💛 Следить"
                elif rsi > 65:
                    signal = "⚠️ Перегрета"
                else:
                    signal = "⚪ Нейтрально"
                results.append((signal, coin, p, ch, rsi))
            except:
                pass
        results.sort(key=lambda x: x[3], reverse=True)
        reply = f"🔍 <b>СКАНЕР МОНЕТ</b>\n<code>══════════════════════</code>\n\n"
        for signal, coin, p, ch, rsi in results[:10]:
            reply += f"{signal} {coin}: ${p:,.2f} | Изм: {ch:+.1f}% | RSI: {rsi}\n"
        reply += f"\n<code>══════════════════════</code>\n<i>RSI < 35 + рост = BUY | RSI > 70 + падение = SELL</i>"
        return reply
    except Exception as e:
        return f"❌ Ошибка сканера: {e}"


def run_backtest(days=30):
    try:
        btc = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=days*24)
        if len(btc) < 48:
            return "⚠️ Недостаточно данных."
        trades = []
        position = None
        entry_price = 0
        for i in range(1, len(btc)):
            close = btc[i][4]
            prev_close = btc[i-1][4]
            volume = btc[i][5]
            prev_volume = btc[i-1][5]
            ch = (close - prev_close) / prev_close * 100
            vol_ch = (volume - prev_volume) / prev_volume * 100 if prev_volume else 0
            
            if position is None:
                # Сигнал BUY: рост >0.5% на объёме
                if ch > 0.5 and volume > prev_volume:
                    position = "LONG"
                    entry_price = close
            elif position == "LONG":
                # Выход: падение >1% или прибыль >3%
                pnl = (close - entry_price) / entry_price * 100
                if pnl > 3 or pnl < -2:
                    trades.append(pnl)
                    position = None
        
        if trades:
            wins = len([t for t in trades if t > 0])
            winrate = wins / len(trades) * 100
            total_pnl = sum(trades)
            avg_pnl = total_pnl / len(trades)
            best = max(trades)
            worst = min(trades)
            return (
                f"📈 <b>БЭКТЕСТ СТРАТЕГИИ ({days} дней)</b>\n"
                f"<code>══════════════════════</code>\n\n"
                f"📋 Сделок: <b>{len(trades)}</b>\n"
                f"✅ Винрейт: <b>{winrate:.1f}%</b>\n"
                f"💰 Общий PnL: <b>{total_pnl:+.2f}%</b>\n"
                f"📊 Средний: <b>{avg_pnl:+.2f}%</b>\n"
                f"🟢 Лучшая: <b>{best:+.2f}%</b>\n"
                f"🔴 Худшая: <b>{worst:+.2f}%</b>\n\n"
                f"<code>══════════════════════</code>"
            )
        return "⚠️ Нет сделок за период."
    except Exception as e:
        return f"❌ Ошибка: {e}"


def ml_predict():
    try:
        from sklearn.linear_model import LinearRegression
        import numpy as np
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=168)  # 7 дней
        if len(ohlcv) < 50:
            return "⚠️ Недостаточно данных для ML."
        closes = [c[4] for c in ohlcv]
        volumes = [c[5] for c in ohlcv]
        X = []
        y = []
        for i in range(24, len(closes)):
            features = []
            for j in range(24):
                features.append(closes[i-24+j])
            features.append(volumes[i])
            X.append(features)
            y.append(closes[i])
        if len(X) < 10:
            return "⚠️ Мало данных."
        X = np.array(X)
        y = np.array(y)
        model = LinearRegression()
        model.fit(X[:-1], y[:-1])
        last_features = np.array([list(closes[-24:]) + [volumes[-1]]])
        pred = model.predict(last_features)[0]
        current = closes[-1]
        change = (pred - current) / current * 100
        if change > 1:
            signal = "🟢 РОСТ"
            action = "Возможен вход в лонг."
        elif change < -1:
            signal = "🔴 ПАДЕНИЕ"
            action = "Готовься к сжатию."
        else:
            signal = "⚪ БОКОВИК"
            action = "Без явного сигнала."
        return (
            f"🤖 <b>ML-ПРОГНОЗ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ Текущая: <b>${current:,.2f}</b>\n"
            f"🎯 Прогноз (1ч): <b>${pred:,.2f}</b> ({change:+.2f}%)\n"
            f"{signal}: {action}\n\n"
            f"<i>Модель: LinearRegression, 168 часов истории</i>"
        )
    except Exception as e:
        return f"❌ ML ошибка: {e}"


def get_a_plus_signal():
    try:
        p, v = get_market_data_rest("BTC")
        if not p or not prev_price or not prev_vol:
            return "⏳ Жду данные..."
        ch = (p - prev_price) / prev_price * 100
        vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
        
        # Расчёт силы (Энергия)
        strength = 5
        if abs(ch) > 2: strength += 2
        elif abs(ch) > 1: strength += 1
        if abs(vol_change) > 20: strength += 2
        elif abs(vol_change) > 10: strength += 1
        if ch > 0 and v > prev_vol: strength += 1
        elif ch < 0 and v > prev_vol: strength -= 1
        if strength > 10: strength = 10
        if strength < 1: strength = 1
        
        # A+ критерии
        checks = []
        checks.append(("Энергия ≥7", strength >= 7))
        checks.append(("Рост >0.5%", ch > 0.5))
        checks.append(("Объём растёт", v > prev_vol))
        checks.append(("Фаза = Эманация", "Эманация" in last_phase))
        checks.append(("Сенсор = Лёгкость", ch > 0.5 and v > prev_vol))
        checks.append(("Не перегрета (RSI<70)", get_rsi() < 70))
        
        passed = sum(1 for _, ok in checks if ok)
        total = len(checks)
        
        if passed == total:
            grade = "⭐ A+ ИДЕАЛЬНЫЙ ВХОД"
            action = "Можно входить. Размер: 2% депозита. Стоп под уровень."
            emoji = "🟢"
        elif passed >= total - 1:
            grade = "💛 A ХОРОШИЙ ВХОД"
            action = "Можно входить осторожно. Размер: 1%."
            emoji = "💛"
        elif passed >= total - 2:
            grade = "⚪ B СРЕДНИЙ"
            action = "Жди улучшения сигнала."
            emoji = "⚪"
        else:
            grade = "🔴 C СЛАБЫЙ"
            action = "Не входи. Жди A+."
            emoji = "🔴"
        
        reply = f"{emoji} <b>{grade}</b>\n"
        reply += f"<code>══════════════════════</code>\n\n"
        reply += f"₿ BTC: ${p:,.2f} | Изм: {ch:+.2f}% | Объём: {vol_change:+.1f}%\n"
        reply += f"⚡ Энергия: {strength}/10 | 📐 Фаза: {last_phase}\n\n"
        reply += f"<b>Проверки ({passed}/{total}):</b>\n"
        for name, ok in checks:
            reply += f"{'✅' if ok else '❌'} {name}\n"
        reply += f"\n💡 <b>{action}</b>"
        return reply
    except Exception as e:
        return f"❌ Ошибка: {e}"


def check_daily_loss_limit():
    try:
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("SELECT SUM(pnl) FROM trades WHERE pnl IS NOT NULL AND date(timestamp) = date('now')")
        row = c.fetchone()
        conn.close()
        daily_pnl = row[0] if row and row[0] else 0
        if daily_pnl < -30:  # -$30 при депозите $1000 = -3%
            return True, daily_pnl
        return False, daily_pnl
    except:
        return False, 0

daily_loss_blocked = False


def get_mtf_signal():
    try:
        signals = {}
        for tf, minutes in [("5m", 5), ("1h", 60), ("4h", 240)]:
            ohlcv = exchange.fetch_ohlcv("BTC/USDT", tf, limit=2)
            if len(ohlcv) < 2:
                signals[tf] = "⚪ Нет данных"
                continue
            close, prev_close = ohlcv[-1][4], ohlcv[-2][4]
            ch = (close - prev_close) / prev_close * 100
            if ch > 0.3:
                signals[tf] = "🟢 Рост"
            elif ch < -0.3:
                signals[tf] = "🔴 Падение"
            else:
                signals[tf] = "⚪ Боковик"
        
        buy_signals = sum(1 for s in signals.values() if "Рост" in s)
        sell_signals = sum(1 for s in signals.values() if "Падение" in s)
        
        if buy_signals == 3:
            verdict = "🟢 <b>ВСЕ ТФ ВВЕРХ — СИЛЬНЫЙ BUY</b>"
            action = "Размер: 2% депозита. Уверенный вход."
        elif buy_signals == 2:
            verdict = "💛 <b>2/3 ВВЕРХ — BUY</b>"
            action = "Размер: 1% депозита. Осторожный вход."
        elif sell_signals >= 2:
            verdict = "🔴 <b>ВНИЗ — HOLD/SELL</b>"
            action = "Не входи. Проверь стопы."
        else:
            verdict = "⚪ <b>РАЗНОБОЙ — HOLD</b>"
            action = "Жди консенсуса таймфреймов."
        
        p = get_price("BTC")
        reply = (
            f"🕐 <b>МУЛЬТИ-ТАЙМФРЕЙМ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ BTC: ${p:,.2f}" + (f"\n\n" if p else "") +
            f"<b>Сигналы по ТФ:</b>\n"
        )
        for tf, sig in signals.items():
            reply += f"  {tf}: {sig}\n"
        reply += (
            f"\n{verdict}\n"
            f"💡 {action}\n\n"
            f"<code>══════════════════════</code>"
        )
        return reply
    except Exception as e:
        return f"❌ Ошибка MTF: {e}"


def analyze_losses():
    try:
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("SELECT coin, pnl, reason, timestamp FROM trades WHERE pnl < 0 ORDER BY id DESC LIMIT 5")
        rows = c.fetchall()
        conn.close()
        if not rows:
            return "✅ Нет убыточных сделок для анализа."
        reply = f"🔍 <b>АНАЛИЗ ОШИБОК</b>\n<code>══════════════════════</code>\n\n"
        patterns = {}
        for coin, pnl, reason, ts in rows:
            reply += f"🔴 {coin}: <b>{pnl:+.2f}</b> | {reason or '—'} | {ts[:16]}\n"
            # Ищем паттерны
            if reason:
                for word in ["Стоп", "Сигнал", "FOMO", "Паника", "Ручной"]:
                    if word in reason:
                        patterns[word] = patterns.get(word, 0) + 1
        if patterns:
            top = max(patterns, key=patterns.get)
            reply += f"\n<b>⚠️ Главная причина убытков:</b> {top} ({patterns[top]} раз)\n"
            lessons = {
                "Стоп": "Ставь стопы шире. Используй ATR x2.",
                "Сигнал": "Входи только по A+ сигналу. Не форсируй.",
                "FOMO": "Жди подтверждения. Не входи на эмоциях.",
                "Паника": "Доверяй Кодексу. Не продавай на дне.",
                "Ручной": "Доверяй агенту. Ручные входы = убытки.",
            }
            reply += f"📋 <b>Урок:</b> {lessons.get(top, 'Анализируй каждую сделку.')}\n"
        reply += f"\n<code>══════════════════════</code>\n<i>Архитектор, ошибки — это данные для роста.</i>"
        return reply
    except Exception as e:
        return f"❌ Ошибка: {e}"


def optimize_strategy():
    try:
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=720)  # 30 дней
        if len(ohlcv) < 100:
            return "⚠️ Недостаточно данных."
        
        best_result = {"winrate": 0, "trades": 0, "pnl": 0, "entry": 0, "exit": 0}
        
        for entry_th in [0.3, 0.5, 0.7, 1.0]:
            for exit_profit in [2, 3, 4, 5]:
                for exit_loss in [1, 1.5, 2, 2.5]:
                    trades = []
                    position = None
                    entry_price = 0
                    for i in range(1, len(ohlcv)):
                        close = ohlcv[i][4]
                        prev_close = ohlcv[i-1][4]
                        ch = (close - prev_close) / prev_close * 100
                        volume = ohlcv[i][5]
                        prev_volume = ohlcv[i-1][5]
                        
                        if position is None:
                            if ch > entry_th and volume > prev_volume:
                                position = "LONG"
                                entry_price = close
                        elif position == "LONG":
                            pnl = (close - entry_price) / entry_price * 100
                            if pnl > exit_profit or pnl < -exit_loss:
                                trades.append(pnl)
                                position = None
                    
                    if len(trades) >= 5:
                        wins = len([t for t in trades if t > 0])
                        wr = wins / len(trades) * 100
                        total = sum(trades)
                        if wr > best_result["winrate"] or (wr == best_result["winrate"] and total > best_result["pnl"]):
                            best_result = {"winrate": wr, "trades": len(trades), "pnl": total, "entry": entry_th, "exit_profit": exit_profit, "exit_loss": exit_loss}
        
        if best_result["trades"] == 0:
            return "⚠️ Не найдено прибыльных параметров."
        
        return (
            f"⚙ <b>ОПТИМИЗАЦИЯ СТРАТЕГИИ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"🔍 Проверено: 64 комбинации параметров\n\n"
            f"<b>🏆 ЛУЧШИЕ ПАРАМЕТРЫ:</b>\n"
            f"• Вход: рост > <b>{best_result['entry']}%</b> на объёме\n"
            f"• Тейк-профит: <b>+{best_result['exit_profit']}%</b>\n"
            f"• Стоп-лосс: <b>-{best_result['exit_loss']}%</b>\n\n"
            f"<b>📊 РЕЗУЛЬТАТ (30 дней):</b>\n"
            f"• Сделок: <b>{best_result['trades']}</b>\n"
            f"• Винрейт: <b>{best_result['winrate']:.1f}%</b>\n"
            f"• Общий PnL: <b>{best_result['pnl']:+.2f}%</b>\n\n"
            f"<code>══════════════════════</code>\n"
            f"<i>Обновляй раз в неделю для адаптации к рынку.</i>"
        )
    except Exception as e:
        return f"❌ Ошибка оптимизации: {e}"


def train_ml_model():
    try:
        import numpy as np
        from sklearn.ensemble import RandomForestClassifier
        from sklearn.preprocessing import StandardScaler
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return None, None
        X, y = [], []
        for i in range(24, len(ohlcv)-1):
            features = [ohlcv[i-24+j][4] for j in range(24)] + [ohlcv[i][5]]
            X.append(features)
            change = (ohlcv[i+1][4] - ohlcv[i][4]) / ohlcv[i][4] * 100
            y.append("BUY" if change > 0.5 else ("SELL" if change < -0.5 else "HOLD"))
        if len(set(y)) < 2:
            return None, None
        X, y = np.array(X), np.array(y)
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        model = RandomForestClassifier(n_estimators=100, max_depth=10, random_state=42)
        model.fit(X_scaled, y)
        return model, scaler
    except Exception as e:
        try: journal_event("ml_error", str(e))
        except: pass
        return None, None

def ml_predict_signal():
    try:
        import numpy as np
        model, scaler = train_ml_model()
        if model is None:
            return "\u26a0 ML: \u043d\u0435\u0434\u043e\u0441\u0442\u0430\u0442\u043e\u0447\u043d\u043e \u0434\u0430\u043d\u043d\u044b\u0445."
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=25)
        features = [c[4] for c in ohlcv[:-1]] + [ohlcv[-1][5]]
        X_scaled = scaler.transform([features])
        pred = model.predict(X_scaled)[0]
        proba = model.predict_proba(X_scaled)[0]
        conf = max(proba) * 100
        emoji = {"BUY": "\U0001f7e2", "SELL": "\U0001f534", "HOLD": "\u26aa"}
        p = get_price("BTC")
        reply = f"\U0001f916 <b>ML-\u0421\u0418\u0413\u041d\u0410\u041b (Random Forest)</b>\n<code>\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550</code>\n\n\u20bf BTC: ${p:,.2f}\n\n\U0001f3af <b>\u041f\u0440\u043e\u0433\u043d\u043e\u0437:</b> {emoji.get(pred, '')} {pred} (\u0443\u0432\u0435\u0440\u0435\u043d\u043d\u043e\u0441\u0442\u044c: {conf:.0f}%)\n\n<b>\u0412\u0435\u0440\u043e\u044f\u0442\u043d\u043e\u0441\u0442\u0438:</b>\n"
        for cls, prob in sorted(zip(model.classes_, proba), key=lambda x: x[1], reverse=True):
            reply += f"  {emoji.get(cls, '')} {cls}: {prob*100:.0f}%\n"
        reply += f"\n<i>\u041c\u043e\u0434\u0435\u043b\u044c: Random Forest, 500 \u0441\u0432\u0435\u0447\u0435\u0439</i>"
        return reply
    except Exception as e:
        return f"\u274c ML \u043e\u0448\u0438\u0431\u043a\u0430: {e}"


def backup_databases():
    try:
        import shutil
        backup_dir = "backups"
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        today = datetime.now().strftime("%Y-%m-%d")
        for db in ["paper_trades.db", "journal.db"]:
            if os.path.exists(db):
                backup_name = f"{backup_dir}/{db.replace('.db', '')}_{today}.db"
                shutil.copy2(db, backup_name)
        # Удаляем старые бэкапы (старше 7 дней)
        cutoff = datetime.now() - timedelta(days=7)
        for f in os.listdir(backup_dir):
            path = os.path.join(backup_dir, f)
            if os.path.isfile(path):
                ftime = datetime.fromtimestamp(os.path.getmtime(path))
                if ftime < cutoff:
                    os.remove(path)
        journal_event("backup", f"Databases backed up ({today})")
    except Exception as e:
        journal_event("backup_error", str(e))


def check_black_swan():
    try:
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=2)
        if len(ohlcv) < 2:
            return
        close, prev = ohlcv[-1][4], ohlcv[-2][4]
        change = (close - prev) / prev * 100
        if change <= -5:
            msg = (
                f"🦢 <b>ЧЁРНЫЙ ЛЕБЕДЬ!</b>\n"
                f"<code>══════════════════════</code>\n\n"
                f"🔴 BTC упал на <b>{change:.1f}%</b> за час!\n"
                f"₿ Цена: <b>${close:,.2f}</b>\n\n"
                f"📋 <b>Кодекс:</b>\n"
                f"• Проверь стоп-лоссы\n"
                f"• Не паникуй\n"
                f"• Не усредняй\n"
                f"• Жди стабилизации\n\n"
                f"<code>══════════════════════</code>"
            )
            send_tg(msg)
            journal_event("black_swan", f"BTC dropped {change:.1f}% in 1h")
    except:
        pass


def calculate_position(symbol, risk_usd, entry=None, stop=None):
    try:
        if not entry:
            ticker = exchange.fetch_ticker(f"{symbol}/USDT")
            entry = ticker["last"]
        if not stop:
            atr = get_atr(f"{symbol}-USDT")
            stop = entry - atr * 2
        risk_per_coin = abs(entry - stop)
        if risk_per_coin == 0:
            return None
        amount = risk_usd / risk_per_coin
        position_value = amount * entry
        return {
            "symbol": symbol,
            "entry": entry,
            "stop": stop,
            "risk_per_coin": risk_per_coin,
            "amount": amount,
            "value": position_value,
            "risk_pct": risk_usd / BALANCE * 100 if BALANCE else 0
        }
    except Exception as e:
        return None


def export_to_csv():
    try:
        import csv
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("SELECT * FROM trades ORDER BY id")
        rows = c.fetchall()
        conn.close()
        if not rows:
            send_tg("📋 Нет сделок для экспорта.")
            return
        filename = f"trades_full_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Timestamp", "Coin", "Action", "Price", "Amount", "PnL", "Reason", "Stop Loss", "Take Profit"])
            for row in rows:
                writer.writerow(list(row) + ["", ""])
        with open(filename, "rb") as f:
            requests.post(f"https://api.telegram.org/bot{TOKEN}/sendDocument",
                          data={"chat_id": CHAT_ID},
                          files={"document": (filename, f, "text/csv")})
        send_tg(f"📋 <b>CSV ЭКСПОРТ</b>\n<code>══════════════════════</code>\n\n✅ Файл: {filename}\n📊 Сделок: {len(rows)}\n\n<i>Открывается в Excel / Google Таблицах</i>")
    except Exception as e:
        send_tg(f"❌ Ошибка CSV: {e}")


def compare_indices():
    try:
        import yfinance as yf
        result = []
        for symbol, name in [("^GSPC", "S&P500"), ("GC=F", "Золото"), ("DX-Y.NYB", "DXY")]:
            try:
                ticker = yf.Ticker(symbol)
                hist = ticker.history(period="1d")
                if not hist.empty:
                    ch = (hist["Close"].iloc[-1] - hist["Open"].iloc[0]) / hist["Open"].iloc[0] * 100
                    result.append(f"{'🟢' if ch>=0 else '🔴'} {name}: {ch:+.2f}%")
                else:
                    result.append(f"⚪ {name}: нет данных")
            except:
                result.append(f"⚪ {name}: ошибка")
        btc = get_price("BTC")
        btc_ch = 0
        if btc and prev_price:
            btc_ch = (btc - prev_price) / prev_price * 100
        reply = (
            f"📊 <b>СРАВНЕНИЕ С ИНДЕКСАМИ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ BTC: ${btc:,.2f} ({btc_ch:+.2f}%)\n\n"
        )
        for r in result:
            reply += f"{r}\n"
        reply += (
            f"\n<code>══════════════════════</code>\n"
            f"<i>Корреляция BTC с индексами</i>"
        )
        return reply
    except Exception as e:
        return f"❌ Ошибка: {e}\n\nУстанови yfinance: pip3 install yfinance"


def show_watchlist():
    try:
        reply = f"📋 <b>ВОТЧЛИСТ</b>\n<code>══════════════════════</code>\n\n"
        for coin in WATCHLIST:
            try:
                ticker = exchange.fetch_ticker(f"{coin}/USDT")
                p = ticker["last"]
                ch = ticker.get("change", 0)
                rsi = get_rsi(f"{coin}-USDT")
                emoji = "🟢" if ch > 0 else ("🔴" if ch < 0 else "⚪")
                reply += f"{emoji} <b>{coin}</b>: ${p:,.2f} ({ch:+.2f}%) | RSI: {rsi}\n"
            except:
                reply += f"⚪ {coin}: нет данных\n"
        reply += f"\n<code>══════════════════════</code>\n<i>/watchlist add BTC | /watchlist remove BTC</i>"
        return reply
    except:
        return "❌ Ошибка вотчлиста."


def get_social_sentiment():
    try:
        # Используем Axiom API (бесплатный, без ключа)
        url = "https://api.alternative.me/fng/?limit=1"
        r = requests.get(url, timeout=10).json()
        fg = r["data"][0]
        fg_value = int(fg["value"])
        fg_text = fg["value_classification"]
        
        # Дополнительно: тренды из Google (бесплатный прокси)
        trends = ""
        try:
            # Пытаемся получить данные о трендах
            r_trends = requests.get("https://api.coingecko.com/api/v3/search/trending", timeout=10).json()
            coins = [c["item"]["name"] for c in r_trends.get("coins", [])[:5]]
            if coins:
                trends = "\n📈 <b>Тренды CoinGecko:</b>\n" + ", ".join(coins)
        except:
            pass
        
        if fg_value <= 25:
            sentiment = "🔴 Страх — толпа паникует. Лучшее время для покупок."
        elif fg_value <= 45:
            sentiment = "🟡 Осторожность — рынок неуверен."
        elif fg_value <= 55:
            sentiment = "⚪ Нейтрально — без эмоций."
        elif fg_value <= 75:
            sentiment = "🟡 Жадность — толпа покупает."
        else:
            sentiment = "🟢 Экстремальная жадность — готовься к коррекции."
        
        reply = (
            f"💬 <b>СЕНТИМЕНТ СОЦСЕТЕЙ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"😱 <b>Страх и Жадность:</b> {fg_value}/100\n"
            f"📊 {fg_text}\n\n"
            f"{sentiment}"
            f"{trends}\n\n"
            f"<code>══════════════════════</code>\n"
            f"<i>Данные: Alternative.me + CoinGecko</i>"
        )
        return reply
    except Exception as e:
        return f"❌ Ошибка: {e}"


def train_lstm_model():
    try:
        import torch
        import torch.nn as nn
        import numpy as np
        from sklearn.preprocessing import StandardScaler
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return None, None
        
        closes = np.array([c[4] for c in ohlcv])
        volumes = np.array([c[5] for c in ohlcv])
        
        # Создаём признаки
        X, y = [], []
        seq_len = 24
        for i in range(seq_len, len(closes)-1):
            features = np.column_stack([closes[i-seq_len:i], volumes[i-seq_len:i]])
            X.append(features)
            future = closes[i+1]
            current = closes[i]
            y.append(1 if future > current * 1.005 else (0 if future < current * 0.995 else 2))
        
        X = np.array(X)
        y = np.array(y)
        
        if len(set(y)) < 2:
            return None, None
        
        # Нормализация
        X_flat = X.reshape(-1, 2)
        scaler = StandardScaler()
        X_flat = scaler.fit_transform(X_flat)
        X = X_flat.reshape(X.shape[0], seq_len, 2)
        
        # LSTM модель
        class LSTMModel(nn.Module):
            def __init__(self):
                super().__init__()
                self.lstm = nn.LSTM(2, 32, 2, batch_first=True)
                self.fc = nn.Linear(32, 3)
            
            def forward(self, x):
                out, _ = self.lstm(x)
                return self.fc(out[:, -1, :])
        
        model = LSTMModel()
        X_tensor = torch.FloatTensor(X)
        y_tensor = torch.LongTensor(y)
        
        # Обучение
        optimizer = torch.optim.Adam(model.parameters(), lr=0.001)
        criterion = nn.CrossEntropyLoss()
        
        for epoch in range(50):
            optimizer.zero_grad()
            output = model(X_tensor)
            loss = criterion(output, y_tensor)
            loss.backward()
            optimizer.step()
        
        return model, scaler, seq_len
    except Exception as e:
        try: journal_event("lstm_error", str(e))
        except: pass
        return None, None, None

def lstm_predict():
    try:
        import torch
        import numpy as np
        
        model, scaler, seq_len = train_lstm_model()
        if model is None:
            return "⚠️ LSTM: недостаточно данных для обучения."
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=seq_len+1)
        closes = np.array([c[4] for c in ohlcv])
        volumes = np.array([c[5] for c in ohlcv])
        
        features = np.column_stack([closes[-seq_len:], volumes[-seq_len:]])
        features = scaler.transform(features)
        X = torch.FloatTensor(features).unsqueeze(0)
        
        model.eval()
        with torch.no_grad():
            output = model(X)
            probs = torch.softmax(output, dim=1)[0]
            pred = torch.argmax(output, dim=1).item()
        
        labels = {0: "🔴 SELL", 1: "🟢 BUY", 2: "⚪ HOLD"}
        p = get_price("BTC")
        
        reply = (
            f"🧠 <b>LSTM-НЕЙРОСЕТЬ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ BTC: ${p:,.2f}" + (f"\n\n" if p else "") +
            f"🎯 <b>Прогноз:</b> {labels.get(pred, '—')}\n\n"
            f"<b>Вероятности:</b>\n"
            f"  🟢 BUY: {probs[1]*100:.0f}%\n"
            f"  🔴 SELL: {probs[0]*100:.0f}%\n"
            f"  ⚪ HOLD: {probs[2]*100:.0f}%\n\n"
            f"<i>Модель: LSTM, 50 эпох, 500 свечей</i>"
        )
        return reply
    except Exception as e:
        return f"❌ LSTM ошибка: {e}"


# Веса Промптов (обновляются автоматически)
PROMPT_WEIGHTS = {
    "sensor": 1.0,
    "blueprint": 1.0,
    "predict": 1.0,
    "energy": 1.0,
    "shadow": 1.0,
    "breath": 1.0,
    "pulse": 1.0,
    "levels": 1.0,
    "compass": 1.0,
    "ml_signal": 1.5,
    "lstm": 2.0,
}

def update_prompt_weights():
    try:
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("SELECT reason, pnl FROM trades WHERE pnl IS NOT NULL AND reason IS NOT NULL ORDER BY id DESC LIMIT 50")
        rows = c.fetchall()
        conn.close()
        if len(rows) < 5:
            return
        stats = {}
        for reason, pnl in rows:
            for prompt in PROMPT_WEIGHTS:
                if prompt in reason.lower():
                    if prompt not in stats:
                        stats[prompt] = {"wins": 0, "total": 0, "pnl": 0}
                    stats[prompt]["total"] += 1
                    stats[prompt]["pnl"] += pnl
                    if pnl > 0:
                        stats[prompt]["wins"] += 1
        for prompt, data in stats.items():
            if data["total"] >= 3:
                winrate = data["wins"] / data["total"] * 100
                avg_pnl = data["pnl"] / data["total"]
                # Новый вес = винрейт/50 + средний PnL/10
                new_weight = winrate / 50 + avg_pnl / 10
                new_weight = max(0.5, min(3.0, new_weight))
                PROMPT_WEIGHTS[prompt] = round(new_weight, 2)
        journal_event("weights", str(PROMPT_WEIGHTS))
    except:
        pass

def show_prompt_weights():
    try:
        update_prompt_weights()
        reply = f"⚖ <b>ВЕСА ПРОМПТОВ</b>\n<code>══════════════════════</code>\n\n"
        sorted_weights = sorted(PROMPT_WEIGHTS.items(), key=lambda x: x[1], reverse=True)
        for name, weight in sorted_weights:
            stars = "⭐" * min(int(weight * 2), 5)
            reply += f"{stars} {name}: <b>{weight:.2f}</b>\n"
        reply += f"\n<code>══════════════════════</code>\n<i>Веса обновляются автоматически по истории сделок.</i>"
        return reply
    except:
        return "❌ Ошибка."


def train_xgboost_model():
    try:
        import xgboost as xgb
        import numpy as np
        from sklearn.preprocessing import StandardScaler
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return None, None
        
        X, y = [], []
        for i in range(24, len(ohlcv)-1):
            features = [ohlcv[i-24+j][4] for j in range(24)] + [ohlcv[i][5]]
            X.append(features)
            change = (ohlcv[i+1][4] - ohlcv[i][4]) / ohlcv[i][4] * 100
            y.append(0 if change > 0.5 else (1 if change < -0.5 else 2))
        
        X, y = np.array(X), np.array(y)
        if len(set(y)) < 2:
            return None, None
        
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        model = xgb.XGBClassifier(n_estimators=100, max_depth=6, learning_rate=0.1, use_label_encoder=False, eval_metric='mlogloss')
        model.fit(X_scaled, y)
        return model, scaler
    except:
        return None, None

def ensemble_predict():
    try:
        import numpy as np
        from sklearn.linear_model import LinearRegression
        from sklearn.ensemble import RandomForestClassifier
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return "⚠️ Ансамбль: недостаточно данных."
        
        votes = {"BUY": 0, "SELL": 0, "HOLD": 0}
        details = []
        
        # 1. Random Forest
        try:
            model_rf, scaler_rf = train_ml_model()
            if model_rf:
                features = [ohlcv[-25+j][4] for j in range(24)] + [ohlcv[-1][5]]
                X = scaler_rf.transform([features])
                pred = model_rf.predict(X)[0]
                votes[pred] += 1
                details.append(f"🌲 Random Forest: {pred}")
        except:
            pass
        
        # 2. LSTM
        try:
            model_lstm, scaler_lstm, seq_len = train_lstm_model()
            if model_lstm:
                import torch
                closes = np.array([c[4] for c in ohlcv])
                volumes = np.array([c[5] for c in ohlcv])
                features = np.column_stack([closes[-seq_len:], volumes[-seq_len:]])
                features = scaler_lstm.transform(features)
                X = torch.FloatTensor(features).unsqueeze(0)
                model_lstm.eval()
                with torch.no_grad():
                    pred_idx = torch.argmax(model_lstm(X), dim=1).item()
                labels = {0: "SELL", 1: "BUY", 2: "HOLD"}
                pred = labels.get(pred_idx, "HOLD")
                votes[pred] += 2  # LSTM имеет двойной вес
                details.append(f"🧠 LSTM: {pred} (x2 вес)")
        except:
            pass
        
        # 3. XGBoost
        try:
            model_xgb, scaler_xgb = train_xgboost_model()
            if model_xgb:
                features = [ohlcv[-25+j][4] for j in range(24)] + [ohlcv[-1][5]]
                X = scaler_xgb.transform([features])
                pred_idx = model_xgb.predict(X)[0]
                labels = {0: "BUY", 1: "SELL", 2: "HOLD"}
                pred = labels.get(pred_idx, "HOLD")
                votes[pred] += 1
                details.append(f"⚡ XGBoost: {pred}")
        except:
            pass
        
        # 4. Linear Regression
        try:
            closes = np.array([c[4] for c in ohlcv])
            X_lr = np.arange(len(closes)).reshape(-1, 1)
            y_lr = closes
            lr = LinearRegression()
            lr.fit(X_lr[-100:], y_lr[-100:])
            pred_val = lr.predict([[len(closes)]])[0]
            change = (pred_val - closes[-1]) / closes[-1] * 100
            pred = "BUY" if change > 0.5 else ("SELL" if change < -0.5 else "HOLD")
            votes[pred] += 1
            details.append(f"📈 Linear: {pred} ({change:+.2f}%)")
        except:
            pass
        
        winner = max(votes, key=votes.get)
        total_votes = sum(votes.values())
        confidence = votes[winner] / total_votes * 100 if total_votes > 0 else 0
        
        emoji = {"BUY": "🟢", "SELL": "🔴", "HOLD": "⚪"}
        p = get_price("BTC")
        
        reply = (
            f"🏛 <b>АНСАМБЛЬ НЕЙРОСЕТЕЙ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ BTC: ${p:,.2f}" + (f"\n\n" if p else "") +
            f"<b>Голосование ({total_votes} голосов):</b>\n"
        )
        for d in details:
            reply += f"  {d}\n"
        reply += (
            f"\n🎯 <b>РЕШЕНИЕ:</b> {emoji.get(winner, '')} {winner}\n"
            f"📊 Уверенность: <b>{confidence:.0f}%</b>\n\n"
            f"<code>══════════════════════</code>\n"
            f"<i>4 модели голосуют. Большинство побеждает.</i>"
        )
        return reply
    except Exception as e:
        return f"❌ Ансамбль ошибка: {e}"


def train_transformer_model():
    try:
        import torch
        import torch.nn as nn
        import numpy as np
        from sklearn.preprocessing import StandardScaler
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return None, None, None
        
        closes = np.array([c[4] for c in ohlcv])
        volumes = np.array([c[5] for c in ohlcv])
        
        X, y = [], []
        seq_len = 24
        for i in range(seq_len, len(closes)-1):
            features = np.column_stack([closes[i-seq_len:i], volumes[i-seq_len:i]])
            X.append(features)
            change = (closes[i+1] - closes[i]) / closes[i] * 100
            y.append(0 if change > 0.5 else (1 if change < -0.5 else 2))
        
        X = np.array(X)
        y = np.array(y)
        
        if len(set(y)) < 2:
            return None, None, None
        
        X_flat = X.reshape(-1, 2)
        scaler = StandardScaler()
        X_flat = scaler.fit_transform(X_flat)
        X = X_flat.reshape(X.shape[0], seq_len, 2)
        
        class TransformerModel(nn.Module):
            def __init__(self):
                super().__init__()
                self.embed = nn.Linear(2, 64)
                encoder_layer = nn.TransformerEncoderLayer(d_model=64, nhead=4, batch_first=True)
                self.transformer = nn.TransformerEncoder(encoder_layer, num_layers=2)
                self.fc = nn.Linear(64, 3)
            
            def forward(self, x):
                x = self.embed(x)
                x = self.transformer(x)
                return self.fc(x[:, -1, :])
        
        model = TransformerModel()
        X_tensor = torch.FloatTensor(X)
        y_tensor = torch.LongTensor(y)
        
        optimizer = torch.optim.Adam(model.parameters(), lr=0.0005)
        criterion = nn.CrossEntropyLoss()
        
        for epoch in range(30):
            optimizer.zero_grad()
            output = model(X_tensor)
            loss = criterion(output, y_tensor)
            loss.backward()
            optimizer.step()
        
        return model, scaler, seq_len
    except Exception as e:
        return None, None, None

def transformer_predict():
    try:
        import torch
        import numpy as np
        
        model, scaler, seq_len = train_transformer_model()
        if model is None:
            return "⚠️ Transformer: недостаточно данных."
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=seq_len+1)
        closes = np.array([c[4] for c in ohlcv])
        volumes = np.array([c[5] for c in ohlcv])
        
        features = np.column_stack([closes[-seq_len:], volumes[-seq_len:]])
        features = scaler.transform(features)
        X = torch.FloatTensor(features).unsqueeze(0)
        
        model.eval()
        with torch.no_grad():
            output = model(X)
            probs = torch.softmax(output, dim=1)[0]
            pred = torch.argmax(output, dim=1).item()
        
        labels = {0: "🟢 BUY", 1: "🔴 SELL", 2: "⚪ HOLD"}
        p = get_price("BTC")
        
        reply = (
            f"🔮 <b>TRANSFORMER-НЕЙРОСЕТЬ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ BTC: ${p:,.2f}" + (f"\n\n" if p else "") +
            f"🎯 <b>Прогноз:</b> {labels.get(pred, '—')}\n\n"
            f"<b>Вероятности:</b>\n"
            f"  🟢 BUY: {probs[0]*100:.0f}%\n"
            f"  🔴 SELL: {probs[1]*100:.0f}%\n"
            f"  ⚪ HOLD: {probs[2]*100:.0f}%\n\n"
            f"<i>Модель: Transformer, 30 эпох, внимание ко всем свечам</i>"
        )
        return reply
    except Exception as e:
        return f"❌ Transformer ошибка: {e}"


KARMA = 0

def update_karma(pnl):
    global KARMA
    if pnl > 0:
        KARMA += 1
    else:
        KARMA -= 1

def get_karma_level():
    if KARMA >= 10:
        return "🟢 Просветлённый"
    elif KARMA >= 5:
        return "💛 Опытный"
    elif KARMA >= 0:
        return "⚪ Нейтральный"
    elif KARMA >= -5:
        return "🟠 Рисковый"
    else:
        return "🔴 Тёмный"

def show_karma():
    try:
        conn = sqlite3.connect("paper_trades.db")
        c = conn.cursor()
        c.execute("SELECT pnl FROM trades WHERE pnl IS NOT NULL ORDER BY id")
        rows = c.fetchall()
        conn.close()
        global KARMA
        KARMA = sum(1 if r[0] > 0 else -1 for r in rows)
        level = get_karma_level()
        reply = (
            f"🙏 <b>КАРМА АГЕНТА</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"🎭 Уровень: <b>{level}</b>\n"
            f"🔢 Счёт: <b>{KARMA:+d}</b>\n"
            f"📊 Сделок: <b>{len(rows)}</b>\n\n"
        )
        if KARMA >= 10:
            reply += "💡 Агент в гармонии. Можно доверять сигналам."
        elif KARMA >= 0:
            reply += "💡 Агент учится. Следи за весами Промптов."
        else:
            reply += "💡 Агент в минусе. Проверь стратегию."
        reply += f"\n\n<code>══════════════════════</code>\n<i>Каждая сделка меняет карму.</i>"
        return reply
    except:
        return "❌ Ошибка."


def train_cnn_model():
    try:
        import torch
        import torch.nn as nn
        import numpy as np
        from sklearn.preprocessing import StandardScaler
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return None, None, None
        
        closes = np.array([c[4] for c in ohlcv])
        volumes = np.array([c[5] for c in ohlcv])
        
        X, y = [], []
        seq_len = 24
        for i in range(seq_len, len(closes)-1):
            features = np.column_stack([closes[i-seq_len:i], volumes[i-seq_len:i]])
            X.append(features)
            change = (closes[i+1] - closes[i]) / closes[i] * 100
            y.append(0 if change > 0.5 else (1 if change < -0.5 else 2))
        
        X = np.array(X)
        y = np.array(y)
        
        if len(set(y)) < 2:
            return None, None, None
        
        X_flat = X.reshape(-1, 2)
        scaler = StandardScaler()
        X_flat = scaler.fit_transform(X_flat)
        X = X_flat.reshape(X.shape[0], 1, seq_len, 2)  # [batch, channels, height, width]
        
        class CNNModel(nn.Module):
            def __init__(self):
                super().__init__()
                self.conv1 = nn.Conv2d(1, 16, kernel_size=(3, 2), padding=(1, 0))
                self.conv2 = nn.Conv2d(16, 32, kernel_size=(3, 1), padding=(1, 0))
                self.pool = nn.AdaptiveAvgPool2d((4, 1))
                self.fc = nn.Linear(32 * 4, 3)
            
            def forward(self, x):
                x = torch.relu(self.conv1(x))
                x = torch.relu(self.conv2(x))
                x = self.pool(x)
                x = x.view(x.size(0), -1)
                return self.fc(x)
        
        model = CNNModel()
        X_tensor = torch.FloatTensor(X)
        y_tensor = torch.LongTensor(y)
        
        optimizer = torch.optim.Adam(model.parameters(), lr=0.001)
        criterion = nn.CrossEntropyLoss()
        
        for epoch in range(30):
            optimizer.zero_grad()
            output = model(X_tensor)
            loss = criterion(output, y_tensor)
            loss.backward()
            optimizer.step()
        
        return model, scaler, seq_len
    except:
        return None, None, None

def cnn_predict():
    try:
        import torch
        import numpy as np
        
        model, scaler, seq_len = train_cnn_model()
        if model is None:
            return "⚠️ CNN: недостаточно данных."
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=seq_len+1)
        closes = np.array([c[4] for c in ohlcv])
        volumes = np.array([c[5] for c in ohlcv])
        
        features = np.column_stack([closes[-seq_len:], volumes[-seq_len:]])
        features = scaler.transform(features)
        X = torch.FloatTensor(features).unsqueeze(0).unsqueeze(0)  # [1, 1, 24, 2]
        
        model.eval()
        with torch.no_grad():
            output = model(X)
            probs = torch.softmax(output, dim=1)[0]
            pred = torch.argmax(output, dim=1).item()
        
        labels = {0: "🟢 BUY", 1: "🔴 SELL", 2: "⚪ HOLD"}
        p = get_price("BTC")
        
        reply = (
            f"👁 <b>CNN-НЕЙРОСЕТЬ (Паттерны)</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ BTC: ${p:,.2f}" + (f"\n\n" if p else "") +
            f"🎯 <b>Прогноз:</b> {labels.get(pred, '—')}\n\n"
            f"<b>Вероятности:</b>\n"
            f"  🟢 BUY: {probs[0]*100:.0f}%\n"
            f"  🔴 SELL: {probs[1]*100:.0f}%\n"
            f"  ⚪ HOLD: {probs[2]*100:.0f}%\n\n"
            f"<i>Модель: CNN, ищет графические паттерны</i>"
        )
        return reply
    except Exception as e:
        return f"❌ CNN ошибка: {e}"


def train_catboost_model():
    try:
        from catboost import CatBoostClassifier
        import numpy as np
        from sklearn.preprocessing import StandardScaler
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return None, None
        
        X, y = [], []
        for i in range(24, len(ohlcv)-1):
            features = [ohlcv[i-24+j][4] for j in range(24)] + [ohlcv[i][5]]
            X.append(features)
            change = (ohlcv[i+1][4] - ohlcv[i][4]) / ohlcv[i][4] * 100
            y.append(0 if change > 0.5 else (1 if change < -0.5 else 2))
        
        X, y = np.array(X), np.array(y)
        if len(set(y)) < 2:
            return None, None
        
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        model = CatBoostClassifier(iterations=100, depth=6, learning_rate=0.1, verbose=0, random_seed=42)
        model.fit(X_scaled, y)
        return model, scaler
    except:
        return None, None

def train_lightgbm_model():
    try:
        from lightgbm import LGBMClassifier
        import numpy as np
        from sklearn.preprocessing import StandardScaler
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return None, None
        
        X, y = [], []
        for i in range(24, len(ohlcv)-1):
            features = [ohlcv[i-24+j][4] for j in range(24)] + [ohlcv[i][5]]
            X.append(features)
            change = (ohlcv[i+1][4] - ohlcv[i][4]) / ohlcv[i][4] * 100
            y.append(0 if change > 0.5 else (1 if change < -0.5 else 2))
        
        X, y = np.array(X), np.array(y)
        if len(set(y)) < 2:
            return None, None
        
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        model = LGBMClassifier(n_estimators=100, max_depth=6, learning_rate=0.1, verbose=-1, random_state=42)
        model.fit(X_scaled, y)
        return model, scaler
    except:
        return None, None

def mega_ensemble_predict():
    try:
        import numpy as np
        import torch
        from sklearn.linear_model import LinearRegression
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "1h", limit=500)
        if len(ohlcv) < 100:
            return "⚠️ Мега-Ансамбль: недостаточно данных."
        
        votes = {"BUY": 0, "SELL": 0, "HOLD": 0}
        details = []
        
        models = [
            ("🌲 Random Forest", 1, lambda: train_ml_model()),
            ("⚡ XGBoost", 1, lambda: train_xgboost_model()),
            ("🐱 CatBoost", 1, lambda: train_catboost_model()),
            ("💡 LightGBM", 1, lambda: train_lightgbm_model()),
        ]
        
        for name, weight, train_fn in models:
            try:
                result = train_fn()
                if result and result[0]:
                    model, scaler = result
                    features = [ohlcv[-25+j][4] for j in range(24)] + [ohlcv[-1][5]]
                    X = scaler.transform([features])
                    pred_idx = model.predict(X)[0]
                    labels = {0: "BUY", 1: "SELL", 2: "HOLD"}
                    pred = labels.get(pred_idx, "HOLD")
                    votes[pred] += weight
                    details.append(f"{name}: {pred}")
            except:
                pass
        
        # LSTM
        try:
            result = train_lstm_model()
            if result[0]:
                model_lstm, scaler_lstm, seq_len = result
                closes = np.array([c[4] for c in ohlcv])
                volumes = np.array([c[5] for c in ohlcv])
                features = np.column_stack([closes[-seq_len:], volumes[-seq_len:]])
                features = scaler_lstm.transform(features)
                X = torch.FloatTensor(features).unsqueeze(0)
                model_lstm.eval()
                with torch.no_grad():
                    pred_idx = torch.argmax(model_lstm(X), dim=1).item()
                labels = {0: "SELL", 1: "BUY", 2: "HOLD"}
                pred = labels.get(pred_idx, "HOLD")
                votes[pred] += 2
                details.append(f"🧠 LSTM: {pred} (x2)")
        except:
            pass
        
        # Transformer
        try:
            result = train_transformer_model()
            if result[0]:
                model_t, scaler_t, seq_len_t = result
                closes = np.array([c[4] for c in ohlcv])
                volumes = np.array([c[5] for c in ohlcv])
                features = np.column_stack([closes[-seq_len_t:], volumes[-seq_len_t:]])
                features = scaler_t.transform(features)
                X = torch.FloatTensor(features).unsqueeze(0)
                model_t.eval()
                with torch.no_grad():
                    pred_idx = torch.argmax(model_t(X), dim=1).item()
                labels = {0: "BUY", 1: "SELL", 2: "HOLD"}
                pred = labels.get(pred_idx, "HOLD")
                votes[pred] += 2
                details.append(f"🔮 Transformer: {pred} (x2)")
        except:
            pass
        
        # CNN
        try:
            result = train_cnn_model()
            if result[0]:
                model_c, scaler_c, seq_len_c = result
                closes = np.array([c[4] for c in ohlcv])
                volumes = np.array([c[5] for c in ohlcv])
                features = np.column_stack([closes[-seq_len_c:], volumes[-seq_len_c:]])
                features = scaler_c.transform(features)
                X = torch.FloatTensor(features).unsqueeze(0).unsqueeze(0)
                model_c.eval()
                with torch.no_grad():
                    pred_idx = torch.argmax(model_c(X), dim=1).item()
                labels = {0: "BUY", 1: "SELL", 2: "HOLD"}
                pred = labels.get(pred_idx, "HOLD")
                votes[pred] += 1
                details.append(f"👁 CNN: {pred}")
        except:
            pass
        
        # Linear
        try:
            closes = np.array([c[4] for c in ohlcv])
            X_lr = np.arange(len(closes)).reshape(-1, 1)
            lr = LinearRegression()
            lr.fit(X_lr[-100:], closes[-100:])
            pred_val = lr.predict([[len(closes)]])[0]
            change = (pred_val - closes[-1]) / closes[-1] * 100
            pred = "BUY" if change > 0.5 else ("SELL" if change < -0.5 else "HOLD")
            votes[pred] += 1
            details.append(f"📈 Linear: {pred} ({change:+.2f}%)")
        except:
            pass
        
        winner = max(votes, key=votes.get)
        total = sum(votes.values())
        confidence = votes[winner] / total * 100 if total > 0 else 0
        
        emoji = {"BUY": "🟢", "SELL": "🔴", "HOLD": "⚪"}
        p = get_price("BTC")
        
        reply = (
            f"🏛 <b>МЕГА-АНСАМБЛЬ (8 моделей)</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ BTC: ${p:,.2f}" + (f"\n\n" if p else "") +
            f"<b>Голосование ({total} голосов):</b>\n"
        )
        for d in details:
            reply += f"  {d}\n"
        reply += (
            f"\n🎯 <b>РЕШЕНИЕ:</b> {emoji.get(winner, '')} {winner}\n"
            f"📊 Уверенность: <b>{confidence:.0f}%</b>\n\n"
            f"<code>══════════════════════</code>\n"
            f"<i>8 нейросетей голосуют. Большинство побеждает.</i>"
        )
        return reply
    except Exception as e:
        return f"❌ Мега-Ансамбль ошибка: {e}"


def autoencoder_detect():
    try:
        import torch
        import torch.nn as nn
        import numpy as np
        from sklearn.preprocessing import StandardScaler
        
        ohlcv = exchange.fetch_ohlcv("BTC/USDT", "5m", limit=200)
        if len(ohlcv) < 100:
            return "⚠️ AutoEncoder: недостаточно данных."
        
        closes = np.array([c[4] for c in ohlcv])
        volumes = np.array([c[5] for c in ohlcv])
        changes = np.diff(closes) / closes[:-1] * 100
        
        features = np.column_stack([changes[-100:], volumes[-100:]])
        scaler = StandardScaler()
        features = scaler.fit_transform(features)
        
        class AutoEncoder(nn.Module):
            def __init__(self):
                super().__init__()
                self.encoder = nn.Sequential(nn.Linear(2, 8), nn.ReLU(), nn.Linear(8, 2))
                self.decoder = nn.Sequential(nn.Linear(2, 8), nn.ReLU(), nn.Linear(8, 2))
            def forward(self, x):
                return self.decoder(self.encoder(x))
        
        model = AutoEncoder()
        X = torch.FloatTensor(features)
        optimizer = torch.optim.Adam(model.parameters(), lr=0.001)
        criterion = nn.MSELoss()
        
        for epoch in range(50):
            optimizer.zero_grad()
            output = model(X)
            loss = criterion(output, X)
            loss.backward()
            optimizer.step()
        
        model.eval()
        with torch.no_grad():
            reconstructed = model(X)
            errors = torch.mean((X - reconstructed) ** 2, dim=1).numpy()
        
        threshold = np.mean(errors) + 2 * np.std(errors)
        anomalies = errors[-10:] > threshold
        anomaly_count = sum(anomalies)
        
        p = get_price("BTC")
        reply = (
            f"🔍 <b>AUTOENCODER — ДЕТЕКТОР АНОМАЛИЙ</b>\n"
            f"<code>══════════════════════</code>\n\n"
            f"₿ BTC: ${p:,.2f}" + (f"\n\n" if p else "") +
            f"📊 Порог аномалии: <b>{threshold:.4f}</b>\n"
            f"⚠️ Аномалий за 50 мин: <b>{anomaly_count}/10</b>\n\n"
        )
        if anomaly_count >= 3:
            reply += "🔴 <b>ОБНАРУЖЕНА МАНИПУЛЯЦИЯ!</b>\nВозможен ложный пробой или памп.\n📋 Кодекс: не входи. Жди подтверждения."
        elif anomaly_count >= 1:
            reply += "🟡 <b>Повышенная волатильность.</b>\nБудь осторожна."
        else:
            reply += "🟢 <b>Рынок чист.</b>\nМанипуляций не обнаружено."
        reply += f"\n\n<code>══════════════════════</code>\n<i>AutoEncoder видит то, что скрыто от глаз.</i>"
        return reply
    except Exception as e:
        return f"❌ AutoEncoder ошибка: {e}"

def update_trailing_stop():
    for coin, d in ACTIVE_PORTFOLIO.items():
        p = get_price(coin)
        if p:
            pnl = (p - d["entry"]) / d["entry"] * 100
            if pnl >= TRAILING_PERCENT:
                new_stop = p * 0.98
                if new_stop > d["stop"]:
                    d["stop"] = round(new_stop, 2)
                    send_voice(f"Трейлинг-стоп {coin}")
                    send_tg(f"🔄 Трейлинг-стоп {coin}: ${d['stop']:,.2f}")

def check_sharp_move():
    global prev_price
    p = get_price("BTC")
    if p and prev_price:
        ch = (p - prev_price) / prev_price * 100
        if abs(ch) >= 1.0:
            direction = "Рост" if ch > 0 else "Падение"
            send_voice(f"Резкое движение! {direction} {abs(ch):.1f}%")
            send_tg(f"⚡ Резкое движение BTC: {direction} {abs(ch):.2f}%, ${p:,.2f}")

def paper_trade_logic():
    global PAPER_BALANCE, PAPER_POSITIONS
    for coin in ["BTC", "ETH", "SOL"]:
        p = get_price(coin)
        if not p: continue
        if coin not in PAPER_POSITIONS or PAPER_POSITIONS[coin] is None:
            entry_threshold = 64000 if coin == "BTC" else (2000 if coin == "ETH" else 82)
            if p < entry_threshold:
                amount = PAPER_BALANCE * 0.02 / p
                PAPER_POSITIONS[coin] = {"entry": p, "amount": amount, "stop": p * 0.98, "target": p * 1.03}
                log_paper_trade(coin, "BUY", p, amount, reason="Сигнал")
                send_tg(f"📝 Paper BUY {coin} @ ${p:,.2f}")
        else:
            pos = PAPER_POSITIONS[coin]
            if p <= pos["stop"]:
                pnl = pos["amount"] * (p - pos["entry"]); PAPER_BALANCE += pnl
                log_paper_trade(coin, "SELL", p, pos["amount"], pnl, "Стоп")
                send_tg(f"📝 Paper SELL {coin} @ ${p:,.2f} (стоп), PnL: ${pnl:+,.2f}")
                PAPER_POSITIONS[coin] = None
            elif p >= pos["target"]:
                pnl = pos["amount"] * (p - pos["entry"]); PAPER_BALANCE += pnl
                log_paper_trade(coin, "SELL", p, pos["amount"], pnl, "Тейк")
                send_tg(f"📝 Paper SELL {coin} @ ${p:,.2f} (тейк), PnL: ${pnl:+,.2f}")
                PAPER_POSITIONS[coin] = None
            else:
                new_stop = p * 0.98
                if new_stop > pos["stop"]: PAPER_POSITIONS[coin]["stop"] = round(new_stop, 2)

keyboard = {"keyboard": [["👋 Привет", "📊 Статус"], ["😱 Страх", "💀 Ликв"], ["⛓ Ончейн", "📖 Стакан"], ["📰 Новости", "🧭 Компас"], ["⚓ Якорь", "📐 Чертёж"], ["🧠 Сенсор", "🔄 Разворот"], ["💡 Совет", "⚡ Энергия"], ["👁 Тень", "🌬 Дыхание"], ["💓 Пульс", "🗺 Уровни"], ["🪞 Зеркало", "🏮 Маяк"], ["🔮 Прогноз", "📝 Paper"], ["📈 Backtest", "📊 Экспорт"], ["📋 Бэктест", "🤖 ML-Прогноз"], ["⚙ Оптимизация"], ["🛑 Дневной лимит", "🔍 Анализ ошибок"], ["📋 Вотчлист"], ["🧮 Калькулятор"], ["📋 История", "🏆 Топ"], ["📊 Статистика", "📊 Дэшборд"], ["🔗 Ссылка", "📈 График"], ["⏱ Аптайм", "🔍 Сканер"], ["🎯 Удар", "📝 Заметка"], ["⭐ A+ Сигнал", "🕐 Мульти-ТФ"], ["📊 Метрики"], ["📋 Заметки", "⚠️ Профиль"], ["🎭 Сентимент"]], "resize_keyboard": True}
print("Архитектор: агент с полным набором Промптов запущен")
init_paper_db()
migrate_db()
p_hello = get_price("BTC")
hello_msg = f"👋 <b>ДОБРОЕ УТРО, АРХИТЕКТОР!</b>\n\n🏰 Все системы активны.\n📊 Фаза: {last_phase}\n"
if p_hello: hello_msg += f"₿ BTC: ${p_hello:,.2f}\n"
hello_msg += "\n💡 Кодекс работает. Империя под контролем."
send_tg(hello_msg, keyboard)

threading.Thread(target=websocket_thread, daemon=True).start()

p, v = get_market_data_rest("BTC")
if p: prev_price = p; prev_vol = v

last_check = time.time(); last_summary = time.time(); last_trailing = time.time()
last_paper = time.time(); last_sharp_check = time.time(); last_daily = time.time(); last_imperative = time.time()
start_time = time.time()

while True:
    process_updates(); time.sleep(0.5)
    now = time.time()
    if now - last_check >= 300:
        check_black_swan()
        for coin, d in ACTIVE_PORTFOLIO.items():
            p = get_price(coin)
            if p:
                if p >= d["target"]: send_voice(f"Тейк-профит {coin}"); send_tg(f"🎯 ТЕЙК-ПРОФИТ {coin}! ${p:,.2f}")
                elif p <= d["stop"]: send_voice(f"Стоп-лосс {coin}"); send_tg(f"🛑 СТОП-ЛОСС {coin}! ${p:,.2f}")
        check_stop_proximity(); check_target_proximity()
        new_phase = last_phase
        p = get_price("BTC")
        if p and prev_price:
            if p > prev_price: last_phase = "📈 Эманация"
            elif p < prev_price: last_phase = "📉 Сжатие"
            else: last_phase = "📊 Боковик"
            if new_phase != last_phase:
                phase_start_time = datetime.now()
                phase_history.append((datetime.now().strftime("%d.%m.%Y %H:%M"), last_phase))
                if len(phase_history) > 10: phase_history = phase_history[-10:]
                # Уведомление в канал о развороте
                try:
                    p = get_price("BTC")
                    arrow = "📈" if "Эманация" in last_phase else ("📉" if "Сжатие" in last_phase else "📊")
                    send_tg(f"🔄 <b>РАЗВОРОТ РЫНКА</b>\n<code>══════════════════════</code>\n\n{arrow} <b>{last_phase}</b>\n₿ BTC: ${p:,.2f}" + (f"\n\n📋 <b>Кодекс:</b> {chr(10).join(['• Проверь стоп-лоссы','• Не паникуй','• Действуй по Чертёжу'])}" if "Сжатие" in last_phase else f"\n\n📋 <b>Кодекс:</b> {chr(10).join(['• Можно наращивать позиции','• Подтягивай стопы','• Следуй за трендом'])}"), to_channel=True)
                except:
                    pass
        if p: prev_price = p
        check_level_alerts()
        last_check = now
    if now - last_sharp_check >= 300: check_sharp_move(); last_sharp_check = now
    if now - last_trailing >= 600: update_trailing_stop(); last_trailing = now
    if now - last_paper >= 300:
        blocked, daily_pnl = check_daily_loss_limit()
        if blocked:
            journal_event("block", f"Daily loss limit reached: ${daily_pnl:,.2f}")
        else:
            paper_trade_logic()
        last_paper = now
    if now - last_daily >= 86400: daily_channel_summary(); clean_old_logs(); backup_databases(); update_prompt_weights(); last_daily = now
    if now - last_imperative >= 14400:  # 4 часа
        try:
            p, v = get_market_data_rest("BTC")
            if p and prev_price and prev_vol:
                ch = (p - prev_price) / prev_price * 100
                vol_change = (v - prev_vol) / prev_vol * 100 if prev_vol else 0
                if ch > 0.5 and v > prev_vol:
                    decision = "🟢 BUY"
                    size = "2%"
                    reason = "Рост на объёме. Эманация."
                elif ch < -0.5 and v > prev_vol:
                    decision = "🔴 SELL / HOLD"
                    size = "0%"
                    reason = "Падение на объёме. Сжатие."
                else:
                    decision = "⚪ HOLD"
                    size = "1%"
                    reason = "Боковик или слабый сигнал."
                send_tg(f"⏰ <b>АВТО-ИМПЕРАТИВ</b>\n<code>══════════════════════</code>\n\n📐 Фаза: {last_phase}\n₿ BTC: ${p:,.2f}\n\n🎯 <b>{decision}</b>\n💰 Размер: {size}\n📋 {reason}\n\n<code>══════════════════════</code>\n<i>Авто-сигнал каждые 4 часа.</i>")
        except:
            pass
        last_imperative = now
    if now - last_summary >= 3600:
        summary = f"📊 <b>СВОДКА</b>\nФаза: {last_phase}\n"
        for coin, d in PORTFOLIO.items():
            p = get_price(coin)
            if p: summary += f"{'🟢' if (p-d['entry'])/d['entry']*100>=0 else '🔴'} {coin}: ${p:,.2f} ({(p-d['entry'])/d['entry']*100:+.2f}%)\n"
        summary += f"\n💰 Баланс: ${BALANCE:,.2f}"
        send_tg(summary); last_summary = now